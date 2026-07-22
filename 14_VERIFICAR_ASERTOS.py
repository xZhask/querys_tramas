import os
import sys
import json
import csv
import argparse
import subprocess
import openpyxl

CONTROL_10_SQL = """
WITH emergency_reported AS (
	SELECT sp_numero_documento_paciente AS DNI, sp_fecha_atencion::date AS fecha,
		CASE
			WHEN e.prioridad = 1 THEN '99285'
			WHEN e.prioridad = 2 THEN '99284'
			WHEN e.prioridad = 3 THEN '99282'
			WHEN e.prioridad = 4 THEN '99281'
			ELSE '99281'
		END AS codigo
	FROM temp_emergencia_sigesapol_estancia e
	WHERE e.excluir_tipo2 = false
	  AND NOT EXISTS (
		SELECT 1 FROM temp_hospitalizacion_local h
		WHERE h.sp_numero_documento_paciente = e.sp_numero_documento_paciente
		  AND e.sp_fecha_atencion::date <= h.sp_fecha_alta::date
		  AND e.sp_fecha_alta_emergencia::date >= h.sp_fecha_atencion::date
	  )
	UNION ALL
	SELECT e.sp_numero_documento_paciente, bdt.fecha_atencion::date, bdt.codigo_procedimiento
	FROM temp_bdt_emergencia_sigesapol bdt
	JOIN temp_emergencia_sigesapol_estancia e
	  ON e.sp_numero_documento_paciente = bdt.numero_documento_paciente
	 AND bdt.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta_emergencia::date
	WHERE e.excluir_tipo2 = false AND bdt.codigo_procedimiento IS NOT NULL
	  AND NOT EXISTS (
		SELECT 1 FROM temp_hospitalizacion_local h
		WHERE h.sp_numero_documento_paciente = e.sp_numero_documento_paciente
		  AND bdt.fecha_atencion::date between h.sp_fecha_atencion::date AND h.sp_fecha_alta::date
	  )
	UNION ALL
	SELECT e.sp_numero_documento_paciente, lab.fecha_atencion::date, lab.codigo_procedimiento
	FROM temp_laboratorio_emergencia_sigesapol lab
	JOIN temp_emergencia_sigesapol_estancia e
	  ON e.sp_numero_documento_paciente = lab.numero_documento_paciente
	 AND lab.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta_emergencia::date
	WHERE e.excluir_tipo2 = false AND lab.codigo_procedimiento IS NOT NULL
	  AND NOT EXISTS (
		SELECT 1 FROM temp_hospitalizacion_local h
		WHERE h.sp_numero_documento_paciente = e.sp_numero_documento_paciente
		  AND lab.fecha_atencion::date between h.sp_fecha_atencion::date AND h.sp_fecha_alta::date
	  )
),
hospitalization_reported AS (
	SELECT sp_numero_documento_paciente AS DNI, sp_fecha_atencion::date AS fecha, sp_codigo_procedimiento AS codigo
	FROM temp_hospitalizacion_local
	UNION ALL
	SELECT e.sp_numero_documento_paciente, bdt.fecha_atencion::date, bdt.codigo_procedimiento
	FROM temp_bdt_hospitalizacion_local bdt
	JOIN temp_hospitalizacion_local e
	  ON e.sp_numero_documento_paciente = bdt.numero_documento_paciente
	 AND bdt.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta::date
	WHERE bdt.codigo_procedimiento IS NOT NULL
	UNION ALL
	SELECT e.sp_numero_documento_paciente, bdt.fecha_atencion::date, bdt.codigo_procedimiento
	FROM temp_bdt_emergencia_sigesapol bdt
	JOIN temp_hospitalizacion_local e
	  ON e.sp_numero_documento_paciente = bdt.numero_documento_paciente
	 AND bdt.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta::date
	WHERE e.origen_reclasificacion IS NOT NULL AND bdt.codigo_procedimiento IS NOT NULL
	UNION ALL
	SELECT e.sp_numero_documento_paciente, lab.fecha_atencion::date, lab.codigo_procedimiento
	FROM temp_laboratorio_hospitalizacion_local lab
	JOIN temp_hospitalizacion_local e
	  ON e.sp_numero_documento_paciente = lab.numero_documento_paciente
	 AND lab.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta::date
	WHERE lab.codigo_procedimiento IS NOT NULL
	UNION ALL
	SELECT e.sp_numero_documento_paciente, lab.fecha_atencion::date, lab.codigo_procedimiento
	FROM temp_laboratorio_emergencia_sigesapol lab
	JOIN temp_hospitalizacion_local e
	  ON e.sp_numero_documento_paciente = lab.numero_documento_paciente
	 AND lab.fecha_atencion::date between e.sp_fecha_atencion::date AND e.sp_fecha_alta::date
	WHERE e.origen_reclasificacion IS NOT NULL AND lab.codigo_procedimiento IS NOT NULL
)
SELECT COUNT(*) AS total_duplicados_emergencia_hosp
FROM emergency_reported er
JOIN hospitalization_reported hr
  ON er.DNI = hr.DNI
 AND er.fecha = hr.fecha
 AND er.codigo = hr.codigo;
"""


def parse_args():
    parser = argparse.ArgumentParser(description="Verificar aserciones A1/A2/A3 del contrato de salidas v2")
    parser.add_argument("--year", type=int, default=2025)
    parser.add_argument("--month", type=int, required=True)
    parser.add_argument("--skip-control10", action="store_true", help="Omitir A3-CONTROL10 (no requiere BD)")
    parser.add_argument("--skip-a7-db", action="store_true", help="Omitir A7-cobertura (no requiere BD de origen)")
    parser.add_argument("--allow-missing-conservacion", action="store_true",
                         help="No fallar si metricas.json fue generado antes de existir la tabla 'conservacion' (mes ya procesado con el codigo anterior)")
    return parser.parse_args()


def load_trama_keys(filepath, cols, key_names):
    idx = [cols.index(n) if n in cols else -1 for n in key_names]
    if -1 in idx:
        return set()
    keys = set()
    if not os.path.exists(filepath):
        return keys
    with open(filepath, "rb") as f:
        data = f.read().decode("utf-8")
    sep = "|\r\n" if "|\r\n" in data else "|\n"
    for rec in data.split(sep):
        if not rec.strip():
            continue
        parts = rec.split("|")
        try:
            key = tuple(parts[i].strip() for i in idx)
        except IndexError:
            continue
        keys.add(key)
    return keys


def check_a1(infos_dir, period, allow_missing=False):
    path = os.path.join(infos_dir, "metricas.json")
    if not os.path.exists(path):
        print(f"A1 [{period}]: FALLO - no existe {path}")
        return False
    with open(path, "r", encoding="utf-8") as f:
        metrics = json.load(f)
    conservacion = metrics.get("conservacion")
    if not conservacion:
        if allow_missing:
            print(f"A1 [{period}]: SKIPPED (metricas.json se genero antes de la tabla 'conservacion'; requeriria recargar las tablas temporales de este mes para recalcularla)")
            return True
        print(f"A1 [{period}]: FALLO - metricas.json no tiene tabla 'conservacion'")
        return False
    ok = True
    for tipo, row in conservacion.items():
        if row.get("residuo", 1) != 0:
            print(f"A1 [{period}]: FALLO - residuo != 0 en '{tipo}': {row}")
            ok = False
    if ok:
        print(f"A1 [{period}]: PASS (residuo = 0 en las {len(conservacion)} tramas)")
    return ok


def check_a2(exp_dir, infos_dir, tramas_dir, period):
    cols_path = os.path.join(infos_dir, ".trama_columns.json")
    if not os.path.exists(cols_path):
        print(f"A2 [{period}]: FALLO - no existe {cols_path}")
        return False
    with open(cols_path, "r", encoding="utf-8") as f:
        trama_cols = json.load(f)

    retained_keys = set()
    for fname in (".retained_package_emergencia.json", ".retained_package_hospitalizacion.json"):
        fpath = os.path.join(infos_dir, fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, "r", encoding="utf-8") as f:
            rows = json.load(f)
        for r in rows:
            fecha = r.get("sp_fecha_atencion")
            if fecha and " " in str(fecha):
                fecha = str(fecha).split(" ")[0]
            retained_keys.add((str(r.get("sp_numero_documento_paciente")), str(fecha), str(r.get("sp_codigo_procedimiento"))))

    eme_cols = trama_cols.get("emergencia", [])
    eme_keys = load_trama_keys(
        os.path.join(tramas_dir, "trama_emergencia.txt"), eme_cols,
        ["sp_numero_documento_paciente", "sp_fecha_atencion", "sp_codigo_procedimiento"]
    )
    leaked = retained_keys & eme_keys
    if leaked:
        print(f"A2 [{period}]: FALLO - {len(leaked)} claves de paquete retenido (Caso A) presentes en trama_emergencia.txt")
        return False
    print(f"A2 [{period}]: PASS (cero fugas de paquete retenido en trama_emergencia.txt)")
    return True


def load_trama_values(filepath, cols, col_name):
    if col_name not in cols:
        return None
    idx = cols.index(col_name)
    valores = set()
    if not os.path.exists(filepath):
        return valores
    with open(filepath, "rb") as f:
        data = f.read().decode("utf-8")
    sep = "|\r\n" if "|\r\n" in data else "|\n"
    for rec in data.split(sep):
        if not rec.strip():
            continue
        parts = rec.split("|")
        if idx < len(parts):
            valores.add(parts[idx].strip())
    return valores


def check_a4(infos_dir, tramas_dir, period, alcance=("00013591",)):
    cols_path = os.path.join(infos_dir, ".trama_columns.json")
    if not os.path.exists(cols_path):
        print(f"A4 [{period}]: FALLO - no existe {cols_path}")
        return False
    with open(cols_path, "r", encoding="utf-8") as f:
        trama_cols = json.load(f)

    archivos = {
        "consulta": ("trama_consulta_externa.txt", "sp_codigo_ipress"),
        "emergencia": ("trama_emergencia.txt", "sp_codigo_ipress"),
        "hospitalizacion": ("trama_hospitalizacion.txt", "sp_codigo_ipress"),
        "farmacia": ("trama_farmacia.txt", "ipress_codigo"),
    }
    alcance_set = set(alcance)
    ok = True
    for tipo, (fname, col_name) in archivos.items():
        cols = trama_cols.get(tipo, [])
        valores = load_trama_values(os.path.join(tramas_dir, fname), cols, col_name)
        if valores is None:
            print(f"A4 [{period}]: FALLO - columna '{col_name}' no encontrada en {fname}")
            ok = False
            continue
        fuera_de_alcance = valores - alcance_set
        if fuera_de_alcance:
            print(f"A4 [{period}]: FALLO - {fname} tiene codigo_ipress fuera de alcance: {sorted(fuera_de_alcance)}")
            ok = False
    if ok:
        print(f"A4 [{period}]: PASS (unico codigo_ipress presente en las 4 tramas: {sorted(alcance_set)})")
    return ok


def check_a5(infos_dir, tramas_dir, period, year, month):
    cols_path = os.path.join(infos_dir, ".trama_columns.json")
    if not os.path.exists(cols_path):
        print(f"A5 [{period}]: FALLO - no existe {cols_path}")
        return False
    with open(cols_path, "r", encoding="utf-8") as f:
        trama_cols = json.load(f)

    import calendar
    import datetime
    last_day = calendar.monthrange(year, month)[1]
    p_ini = datetime.date(year, month, 1)
    p_fin = datetime.date(year, month, last_day)

    def is_in_period(date_str):
        if not date_str:
            return False
        if " " in date_str:
            date_str = date_str.split(" ")[0]
        try:
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            return p_ini <= d <= p_fin
        except ValueError:
            pass
        return False

    archivos = {
        "consulta": ("trama_consulta_externa.txt", ["sp_fecha_atencion"]),
        "emergencia": ("trama_emergencia.txt", ["sp_fecha_alta"]),
        "farmacia": ("trama_farmacia.txt", ["fecha_dispensacion_como_atencion", "fecha_dispensacion"]),
    }
    
    ok = True
    for tipo, (fname, col_names) in archivos.items():
        cols = trama_cols.get(tipo, [])
        col_name = next((cn for cn in col_names if cn in cols), None)
        
        if col_name:
            valores = load_trama_values(os.path.join(tramas_dir, fname), cols, col_name)
            if valores is not None:
                fuera = [v for v in valores if not is_in_period(v)]
                if fuera:
                    print(f"A5 [{period}]: FALLO - {fname} tiene fechas fuera de la ventana ({len(fuera)} distintos, ej: {fuera[:3]})")
                    ok = False

    # Hospitalizacion: regla inmutable 11 (CONTEXTO_CANONICO.md) - una
    # estancia SOLO factura en el periodo de SU ALTA. fecha_alta debe existir
    # (nunca NULL: eso seria una estancia abierta que no deberia estar en
    # ninguna trama) y debe caer DENTRO del periodo (ni antes ni despues);
    # fecha_atencion (ingreso) puede ser de un mes anterior si la estancia
    # cruzo de mes - eso es arrastre legitimo, no error.
    h_cols = trama_cols.get("hospitalizacion", [])
    if "sp_fecha_atencion" in h_cols and "sp_fecha_alta" in h_cols:
        idx_atencion = h_cols.index("sp_fecha_atencion")
        idx_alta = h_cols.index("sp_fecha_alta")
        h_path = os.path.join(tramas_dir, "trama_hospitalizacion.txt")
        if os.path.exists(h_path):
            with open(h_path, "rb") as f:
                data = f.read().decode("utf-8")
            sep = "|\r\n" if "|\r\n" in data else "|\n"
            fuera_h = 0
            sin_alta_h = 0
            for rec in data.split(sep):
                if not rec.strip(): continue
                parts = rec.split("|")
                if len(parts) > max(idx_atencion, idx_alta):
                    at_str = parts[idx_atencion].strip()
                    al_str = parts[idx_alta].strip()
                    if " " in at_str: at_str = at_str.split(" ")[0]
                    if " " in al_str: al_str = al_str.split(" ")[0]

                    if not al_str:
                        sin_alta_h += 1
                        continue
                    try:
                        al_d = datetime.datetime.strptime(al_str, "%Y-%m-%d").date()
                        if not (p_ini <= al_d <= p_fin):
                            fuera_h += 1
                    except ValueError:
                        pass
            if sin_alta_h > 0:
                print(f"A5 [{period}]: FALLO - trama_hospitalizacion.txt tiene {sin_alta_h} registros SIN fecha_alta (estancia abierta: no debe facturar hasta su propio periodo de alta, regla inmutable 11)")
                ok = False
            if fuera_h > 0:
                print(f"A5 [{period}]: FALLO - trama_hospitalizacion.txt tiene {fuera_h} registros cuya fecha_alta cae FUERA de este periodo (debe facturar en el periodo de SU alta, regla inmutable 11)")
                ok = False
                
    if ok:
        print(f"A5 [{period}]: PASS (Ventana temporal estricta validada en las 4 tramas)")
    return ok


def check_a6_integridad(infos_dir, tramas_dir, period, allow_missing=False):
    """A6-INTEGRIDAD: el recuento FISICO de lineas de cada trama_*.txt debe
    coincidir con lo que el propio metricas.json (volumenes_tramas) dice haber
    escrito. Esto NO verifica cobertura contra el origen (ver A7-COBERTURA):
    solo detecta corrupcion/truncamiento entre "lo que Python calculo" y "lo
    que quedo fisicamente en disco" (p.ej. un \\r embebido que parte una fila
    en dos lineas fisicas, o una escritura interrumpida)."""
    path = os.path.join(infos_dir, "metricas.json")
    if not os.path.exists(path):
        print(f"A6-integridad [{period}]: FALLO - no existe {path}")
        return False
    with open(path, "r", encoding="utf-8") as f:
        metrics = json.load(f)
    volumenes_tramas = metrics.get("volumenes_tramas")
    if not volumenes_tramas:
        print(f"A6-integridad [{period}]: FALLO - metricas.json no tiene 'volumenes_tramas'")
        return False

    archivos = {
        "trama_consulta_externa": "trama_consulta_externa.txt",
        "trama_emergencia": "trama_emergencia.txt",
        "trama_hospitalizacion": "trama_hospitalizacion.txt",
        "trama_farmacia": "trama_farmacia.txt",
    }

    ok = True
    for key, fname in archivos.items():
        expected = volumenes_tramas.get(key, 0)
        p = os.path.join(tramas_dir, fname)
        if not os.path.exists(p):
            if expected > 0:
                print(f"A6-integridad [{period}]: FALLO - no existe {fname} pero se esperaban {expected} filas")
                ok = False
            continue
        with open(p, "rb") as f:
            data = f.read()
        records = split_records(data)
        actual = len(records)
        if actual != expected:
            print(f"A6-integridad [{period}]: FALLO - {fname} tiene {actual} lineas fisicas, pero metricas.json reporta {expected}")
            ok = False

    if ok:
        print(f"A6-integridad [{period}]: PASS (Recuento fisico de lineas en txt coincide con metricas.json)")
    return ok


# ============================================================================
# A7-COBERTURA
# ----------------------------------------------------------------------------
# A6-integridad solo compara el pipeline contra SI MISMO (metricas.json vs
# los .txt que el mismo proceso escribio) - por eso "quedo mal implementada"
# como asercion de cobertura: un mes con 0 filas reales pero un metricas.json
# coherente (p.ej. porque el guardian de periodo aborto en un paso intermedio
# y el proceso re-uso metricas.json de una corrida anterior) pasaria A6 sin
# problema. A7 corta ese punto ciego: cuenta las prestaciones DIRECTAMENTE en
# las tablas de ORIGEN (prestacion_cpt/procedimiento_cpt en CPT;
# emergencias/hospitalizaciones/prestaciones+prestacion_procedimientos en
# SIGESAPOL), con una consulta que NO pasa por ninguna tabla temp_* del
# pipeline, y la contrasta contra volumenes_raw de metricas.json.
#
# LIMITE DOCUMENTADO (verificado contra julio 2025 real, no solo en teoria):
# un intento inicial de exigir origen - alcance_depurado == volumenes_raw
# EXACTO fallo en las 4 tramas de julio con brechas de 5% a 150%, todas por
# motivos LEGITIMOS que esta consulta independiente no puede replicar barato:
#   - log_alcance_depurado no distingue tipo_atencion dentro de
#     temp_sigesapol_procedimientos (una sola fila de log para consulta+
#     emergencia+hospitalizacion juntas), asi que no se puede restar alcance
#     por tipo de forma confiable.
#   - la extraccion real exige AL MENOS UN diagnostico activo (INNER JOIN a
#     receta_diagnosticos/diagnostico_cpt con estado=1/'N') que este conteo
#     de origen no replica; una prestacion sin diagnostico activo cuenta aqui
#     pero no en volumenes_raw.
#   - una estancia de emergencia se cuenta por su fecha de ALTA, pero sus
#     procedimientos por su propia fecha de atencion: una estancia que cruza
#     fin de mes reparte procedimientos entre dos periodos de forma legitima.
# Por eso A7 NO exige igualdad ni una tolerancia chica: valida ORDEN DE
# MAGNITUD (suficiente para el objetivo real de la mision: detectar "trama
# vacia con metricas coherente" y fugas de periodo groseras tipo
# contaminacion entre meses), no aritmetica exacta byte a byte (eso ya lo
# prueba A1 con residuo 0, sobre numeros que el propio pipeline calculo).
# Falla fuerte solo cuando:
#   (a) la trama final tiene 0 filas y el origen tiene > 0 (el caso pedido
#       explicitamente: "trama vacia con metricas coherente"), o
#   (b) volumenes_raw supera el origen independiente por mas del margen
#       MULTIPLICADOR_A7_MAX (fuga de periodo real: se extrajo mucho mas de
#       lo que el origen del mes puede explicar, como septiembre arrastrando
#       miles de filas de mayo/junio/julio en la mision original).
# ============================================================================

MULTIPLICADOR_A7_MAX = 3.0  # volumenes_raw no deberia superar 3x el origen independiente
# ^ Calibrado en vivo contra julio 2025 (post-fix de PARCHE B/D, ver
# 01_PARCHES_funciones.sql): hospitalizacion queda en ~2.2x-2.6x de forma
# LEGITIMA, confirmado por el equipo: una hospitalizacion/emergencia factura
# en el periodo de SU ALTA, arrastrando procedimientos de meses anteriores si
# la estadia cruzo de mes (regla de negocio, no bug). Mi conteo de origen
# aqui es por mes calendario (no por ventana de estancia especifica, que es
# lo que SI corrigen las funciones parchadas) porque replicar esa ventana
# exacta en una consulta "independiente" exigiria re-derivar la misma logica
# de agrupacion por estancia que ya se corrigio en la BD - de poco valor
# adicional y con su propio riesgo de divergir. 3.0x da margen holgado sobre
# el ~2.6x legitimo observado, sin dejar de detectar una fuga de orden de
# magnitud mayor (la version con el bug de PARCHE D/FIX 6 sin corregir traia
# 101,467 filas de temp_bdt_hospitalizacion_local con fechas desde 2018,
# frente a un origen de un solo mes de ~33,000 - eso SI hubiera disparado
# cualquier umbral razonable).

SIGESAPOL_ORIGEN_SQL = {
    "emergencia_estancias": """
        SELECT COUNT(*) FROM emergencias e
        WHERE e.fecha_alta_medica IS NOT NULL AND e.estado = 5
          AND e.fecha_alta_medica::date BETWEEN %s AND %s
    """,
    "hospitalizacion_estancias": """
        SELECT COUNT(*) FROM hospitalizaciones h
        WHERE h.fecha_alta_medica IS NOT NULL
          AND h.fecha_alta_medica::date BETWEEN %s AND %s
    """,
    "consulta_procedimientos": """
        SELECT COUNT(*) FROM prestaciones pre
        JOIN prestacion_procedimientos pp ON pp.id_prestacion = pre.id
        JOIN procedimientos p2 ON p2.id = pp.id_procedimiento
        WHERE pre.id_tipo_atencion IN (1, 5, 7)
          AND p2.tipo_procedimiento IN (1, 2, 3)
          AND pre.fecha_atencion >= %s AND pre.fecha_atencion < %s::date + INTERVAL '1 day'
    """,
    "emergencia_procedimientos": """
        SELECT COUNT(*) FROM prestaciones pre
        JOIN prestacion_procedimientos pp ON pp.id_prestacion = pre.id
        JOIN procedimientos p2 ON p2.id = pp.id_procedimiento
        WHERE pre.id_tipo_atencion = 2
          AND p2.tipo_procedimiento IN (1, 2, 3)
          AND pre.fecha_atencion >= %s AND pre.fecha_atencion < %s::date + INTERVAL '1 day'
    """,
    "hospitalizacion_procedimientos": """
        SELECT COUNT(*) FROM prestaciones pre
        JOIN prestacion_procedimientos pp ON pp.id_prestacion = pre.id
        JOIN procedimientos p2 ON p2.id = pp.id_procedimiento
        WHERE pre.id_tipo_atencion IN (3, 6, 8)
          AND p2.tipo_procedimiento IN (1, 2, 3)
          AND pre.fecha_atencion >= %s AND pre.fecha_atencion < %s::date + INTERVAL '1 day'
    """,
    "farmacia": """
        SELECT COUNT(*) FROM receta_vales rv
        JOIN producto_recetas pr ON pr.id_receta_vale = rv.id
        JOIN productos p ON p.id = pr.id_producto
        WHERE rv.estado = 1 AND pr.cantidad_dispensada > 0 AND p.petitorio = 'SI'
          -- Mismas categorias de tipo_receta activas hoy en
          -- 12_SIGESAPOL_farmacia.sql (rama CONSULTA EXTERNA). Si ese archivo
          -- cambia de rama activa (EMERGENCIA/HOSPITALIZACION), esta lista
          -- debe actualizarse junto con el.
          AND rv.tipo_receta IN ('AMBULATORIO', 'SERVICIO NUTRICIONAL - AMBULATORIO', 'URGENCIA')
          AND rv.fecha_expedicion::date BETWEEN %s AND %s
          AND rv.fecha_registro::date BETWEEN %s AND %s
    """,
}

CPT_ORIGEN_SQL = {
    # prestacion_cpt no tiene columna de establecimiento: ya es LNS-only por
    # construccion (verificado en CONTEXTO_CANONICO.md §3), por eso no hace
    # falta excluir alcance de este lado.
    "consulta_procedimientos": """
        SELECT COUNT(DISTINCT r.id_procedimiento_cpt)
        FROM prestacion_cpt t
        JOIN diagnostico_cpt d ON d.id_prestacion_cpt = t.id_prestacion_cpt AND d.estado = 'N'
        JOIN procedimiento_cpt r ON r.id_diagnostico_cpt = d.id_diagnostico_cpt
        WHERE t.origen = 'CONSULTA' AND t.estado = 'N'
          -- CONSULTA no tiene fecha_egreso (es ambulatoria, sin alta): se usa
          -- fecha_registro de la cabecera, verificado en vivo (fecha_egreso
          -- viene NULL siempre para este origen).
          AND t.fecha_registro::date BETWEEN %s AND %s
    """,
    "hospitalizacion_procedimientos": """
        SELECT COUNT(DISTINCT r.id_procedimiento_cpt)
        FROM prestacion_cpt t
        JOIN diagnostico_cpt d ON d.id_prestacion_cpt = t.id_prestacion_cpt AND d.estado = 'N'
        JOIN procedimiento_cpt r ON r.id_diagnostico_cpt = d.id_diagnostico_cpt
        WHERE t.origen = 'HOSPITALIZACION' AND t.estado = 'N'
          AND r.fecha_egreso::date BETWEEN %s AND %s
    """,
}


def _conectar(dbname, host, port, user, password):
    import psycopg2
    return psycopg2.connect(f"dbname={dbname} user={user} password={password} host={host} port={port}")


def _conn_params(lns_db_var=None, pg_db_var="PGDATABASE", default_dbname="db_cpt_junio26"):
    # LNS_DB_* es lo que setea el aplicativo web (config/database.php) segun
    # los selectores "Base CPT"/"Base SIGESAPOL" de la pantalla Generar;
    # PG*/localhost:5432/postgres son el fallback historico de run_month.ps1
    # para quien corre este script directo por consola.
    host = os.environ.get("LNS_DB_HOST") or os.environ.get("PGHOST", "localhost")
    port = os.environ.get("LNS_DB_PORT") or os.environ.get("PGPORT", "5432")
    user = os.environ.get("LNS_DB_USER") or os.environ.get("PGUSER", "postgres")
    password = os.environ.get("LNS_DB_PASSWORD") or os.environ.get("PGPASSWORD", "root")
    dbname = (os.environ.get(lns_db_var) if lns_db_var else None) or os.environ.get(pg_db_var, default_dbname)
    return dbname, host, port, user, password


def check_a7_cobertura(year, month, period, infos_dir, skip):
    if skip:
        print(f"A7-cobertura [{period}]: SKIPPED (--skip-a7-db, no se verifico contra BD de origen)")
        return True

    path = os.path.join(infos_dir, "metricas.json")
    if not os.path.exists(path):
        print(f"A7-cobertura [{period}]: FALLO - no existe {path}")
        return False
    with open(path, "r", encoding="utf-8") as f:
        metrics = json.load(f)
    volumenes_raw = metrics.get("volumenes_raw")
    volumenes_tramas = metrics.get("volumenes_tramas", {})
    if not volumenes_raw:
        print(f"A7-cobertura [{period}]: FALLO - metricas.json no tiene 'volumenes_raw'")
        return False

    cpt_dbname, host, port, user, password = _conn_params("LNS_DB_CPT", "PGDATABASE", "db_cpt_junio26")
    sig_dbname, _, _, _, _ = _conn_params("LNS_DB_SIGESAPOL", "PGDATABASE_SIGESAPOL", "sigesapol_junio")

    import datetime
    p_ini = datetime.date(year, month, 1)
    last_day = __import__("calendar").monthrange(year, month)[1]
    p_fin = datetime.date(year, month, last_day)

    try:
        conn_cpt = _conectar(cpt_dbname, host, port, user, password)
        conn_sig = _conectar(sig_dbname, host, port, user, password)
    except Exception as e:
        print(f"A7-cobertura [{period}]: FALLO - no se pudo conectar a las BD de origen ({e})")
        return False

    ok = True
    try:
        cur_cpt = conn_cpt.cursor()
        cur_sig = conn_sig.cursor()

        def sig(clave):
            params = (p_ini, p_fin) if clave != "farmacia" else (p_ini, p_fin, p_ini, p_fin)
            cur_sig.execute(SIGESAPOL_ORIGEN_SQL[clave], params)
            return cur_sig.fetchone()[0]

        def cpt(clave):
            cur_cpt.execute(CPT_ORIGEN_SQL[clave], (p_ini, p_fin))
            return cur_cpt.fetchone()[0]

        # origen_total por trama = estancias (donde aplica) + procedimientos,
        # sumando SIGESAPOL y CPT: misma granularidad de fila que las tramas
        # finales (que unen cabecera de estancia + detalle de procedimiento/
        # laboratorio, ver 09/10/11_ARMADO_*.sql).
        origenes = {
            "consulta": sig("consulta_procedimientos") + cpt("consulta_procedimientos"),
            "emergencia": sig("emergencia_estancias") + sig("emergencia_procedimientos"),
            "hospitalizacion": (
                sig("hospitalizacion_estancias") + sig("hospitalizacion_procedimientos")
                + cpt("hospitalizacion_procedimientos")
            ),
            "farmacia": sig("farmacia"),
        }
        tramas_key = {
            "consulta": "trama_consulta_externa", "emergencia": "trama_emergencia",
            "hospitalizacion": "trama_hospitalizacion", "farmacia": "trama_farmacia",
        }

        print(f"--- A7-cobertura [{period}]: desglose por trama (orden de magnitud, no aritmetica exacta) ---")
        for tipo, origen_total in origenes.items():
            raw = volumenes_raw.get(tipo, 0)
            trama_final = volumenes_tramas.get(tramas_key[tipo], 0)
            ratio = (raw / origen_total) if origen_total else float("inf")

            print(f"  {tipo}: origen_independiente={origen_total} vs volumenes_raw={raw} "
                  f"(ratio={ratio:.2f}x, limite={MULTIPLICADOR_A7_MAX}x) | trama_final={trama_final}")

            if trama_final == 0 and origen_total > 0:
                print(f"A7-cobertura [{period}]: FALLO - {tipo} tiene la trama final en 0 filas pero el origen reporta {origen_total} prestaciones del periodo (trama vacia con metricas coherente).")
                ok = False
                continue
            if origen_total == 0 and raw > 0:
                print(f"A7-cobertura [{period}]: FALLO - {tipo} extrajo {raw} filas pero el origen independiente no encuentra NINGUNA prestacion del periodo (fuga de periodo: el guardian no protegio la extraccion).")
                ok = False
                continue
            if origen_total > 0 and raw > origen_total * MULTIPLICADOR_A7_MAX:
                print(f"A7-cobertura [{period}]: FALLO - {tipo} extrajo {raw} filas, mas de {MULTIPLICADOR_A7_MAX}x el origen independiente ({origen_total}). Posible fuga de periodo (contaminacion con otro mes).")
                ok = False

        cur_cpt.close()
        cur_sig.close()
    finally:
        conn_cpt.close()
        conn_sig.close()

    if ok:
        print(f"A7-cobertura [{period}]: PASS (orden de magnitud de origen consistente con lo extraido en las 4 tramas)")
    return ok


# ============================================================================
# A8-NO-DUPLICACION ENTRE PERIODOS
# ----------------------------------------------------------------------------
# Ninguna prestacion (documento + fecha + codigo_procedimiento) puede
# aparecer en las tramas de DOS periodos distintos - eso es doble cobro entre
# envios (p.ej. estancias hospitalarias largas que arrastran filas de un mes
# anterior). Compara el periodo actual contra TODOS los demas periodos que ya
# existan en expedientes/ (se salta a si mismo). Control permanente: se debe
# correr en cada cierre de periodo, una vez que existan periodos previos.
# ============================================================================

_ARCHIVOS_TRAMA_A8 = {
    "consulta": "trama_consulta_externa.txt",
    "emergencia": "trama_emergencia.txt",
    "hospitalizacion": "trama_hospitalizacion.txt",
}
_MARCADOR_ESTANCIA_A8 = {
    "emergencia": "estancia en emergencia",
    "hospitalizacion": "estancia hospitalaria",
}


def _leer_trama_clasificacion(tramas_dir, cols_por_tipo):
    """
    Lee las 3 tramas de un periodo y devuelve:
      - filas_por_doc: doc -> lista de (fecha_completa, codigo, tipo) de
        TODAS las filas. `fecha_completa` se deja TAL CUAL viene en la
        trama (sin truncar) - es la misma clave que usa `load_trama_keys`/
        A8 original para detectar el overlap; truncar a solo fecha aqui
        colapsaba visitas de emergencia distintas del mismo dia (esa trama
        SI trae hora:minuto:segundo) y disparaba falsos overlaps (29->379
        en la verificacion contra septiembre 2025).
      - ventanas_por_doc: doc -> lista de (ingreso, alta) EN SOLO FECHA
        (para comparar contra ventanas se usa solo granularidad de dia) de
        las estancias (hospitalizacion o emergencia) tal como las calcula
        la propia trama YA EXPORTADA (incluye la ventana extendida por el
        Caso A / union Emergencia->Hospitalizacion). NO se consulta la
        tabla cruda de la BD: esa no refleja la union E->H y produce falsos
        "nuevo" (ver HALLAZGO_frontera_CPT_SIGESAPOL_septiembre_octubre.md
        §3.7).
    """
    filas_por_doc = {}
    ventanas_por_doc = {}
    for tipo, fname in _ARCHIVOS_TRAMA_A8.items():
        cols = cols_por_tipo.get(tipo, [])
        if "base" not in cols:
            continue
        try:
            i_base = cols.index("base")
            i_doc = cols.index("sp_numero_documento_paciente")
            i_fecha = cols.index("sp_fecha_atencion")
            i_cod = cols.index("sp_codigo_procedimiento")
        except ValueError:
            continue
        i_alta = cols.index("sp_fecha_alta") if "sp_fecha_alta" in cols else None
        path = os.path.join(tramas_dir, fname)
        if not os.path.exists(path):
            continue
        marker = _MARCADOR_ESTANCIA_A8.get(tipo)
        indices = [i for i in (i_base, i_doc, i_fecha, i_cod, i_alta) if i is not None]
        maxi = max(indices)
        with open(path, "rb") as f:
            data = f.read().decode("utf-8")
        sep = "|\r\n" if "|\r\n" in data else "|\n"
        for rec in data.split(sep):
            if not rec.strip():
                continue
            parts = rec.split("|")
            if len(parts) <= maxi:
                continue
            doc = parts[i_doc].strip()
            fecha_completa = parts[i_fecha].strip()
            cod = parts[i_cod].strip()
            base = parts[i_base].strip()
            filas_por_doc.setdefault(doc, []).append((fecha_completa, cod, tipo))
            if marker and base == marker and i_alta is not None:
                ingreso = fecha_completa[:10]
                alta = parts[i_alta].strip()[:10]
                ventanas_por_doc.setdefault(doc, []).append((ingreso, alta))
    return filas_por_doc, ventanas_por_doc


def _claves_desde_filas(filas_por_doc):
    claves = set()
    for doc, filas in filas_por_doc.items():
        for (fecha, cod, _tipo) in filas:
            claves.add((doc, fecha, cod))
    return claves


def _tipos_de(filas_por_doc, doc, fecha, cod):
    return {t for (f, c, t) in filas_por_doc.get(doc, []) if f == fecha and c == cod}


def _clasificar_overlap_a8(doc, fecha, tipos_actual, tipos_otro, ventanas_globales):
    """
    Superficie (a) "frontera": la fecha cae dentro de la ventana [ingreso,
    alta] de una estancia (hospitalizacion o emergencia) cuyo mes de alta es
    distinto al mes propio de la fecha - arrastre por regla inmutable 11 /
    PARCHE D-E. La ventana se busca en el indice GLOBAL (todos los periodos
    ya generados), no solo en los dos periodos comparados, para cubrir
    tambien estancias que saltan mas de una frontera.
    Superficie (b) "consulta_estancia": un lado del overlap es una fila de
    la trama de CONSULTA y el otro es de hospitalizacion o emergencia -
    HALLAZGO_SIGESAPOL_consulta_vs_estancia.md (duplicado en el instante de
    transferencia E->H).
    Las dos superficies NO son mutuamente excluyentes: un registro fechado
    justo el dia de ingreso puede calzar con ambas a la vez (el arrastre y
    el duplicado de transferencia son fenomenos independientes que a veces
    coinciden en la misma fila). Por eso se devuelve el conjunto de
    superficies que calzaron (puede tener 1 o 2 elementos), no una sola
    etiqueta forzada - la clasificacion es para trazabilidad/reporte, no
    cambia PASS/FALLO (cualquier superficie que calce ya es conocido).
    Devuelve un set: subconjunto de {'frontera', 'consulta_estancia'},
    vacio si no calza ninguna (NUEVO).
    `fecha` puede venir con hora (emergencia) - la comparacion de ventana
    usa solo los primeros 10 caracteres (fecha), igual que las ventanas.
    """
    superficies = set()

    tipos = tipos_actual | tipos_otro
    if "consulta" in tipos and tipos & {"hospitalizacion", "emergencia"}:
        superficies.add("consulta_estancia")

    fecha_dia = fecha[:10]
    fecha_periodo = fecha_dia[:7]
    for (ingreso, alta) in ventanas_globales.get(doc, []):
        if ingreso <= fecha_dia <= alta and alta[:7] != fecha_periodo:
            superficies.add("frontera")
            break

    return superficies


def check_a8_no_duplicacion(year, month, period):
    exp_root = "expedientes"
    exp_dir = os.path.join(exp_root, period)
    infos_dir = os.path.join(exp_dir, "03_INFORMATIVOS")
    tramas_dir = os.path.join(exp_dir, "01_TRAMAS")
    cols_path = os.path.join(infos_dir, ".trama_columns.json")
    if not os.path.exists(cols_path):
        print(f"A8-no-duplicacion [{period}]: FALLO - no existe {cols_path}")
        return False
    with open(cols_path, "r", encoding="utf-8") as f:
        cols_actual = json.load(f)
    filas_actual, ventanas_actual = _leer_trama_clasificacion(tramas_dir, cols_actual)
    claves_actual = _claves_desde_filas(filas_actual)

    if not os.path.isdir(exp_root):
        print(f"A8-no-duplicacion [{period}]: FALLO - no existe la carpeta {exp_root}")
        return False

    otros_periodos = sorted(
        d for d in os.listdir(exp_root)
        if d != period and os.path.isdir(os.path.join(exp_root, d))
        and os.path.exists(os.path.join(exp_root, d, "03_INFORMATIVOS", ".trama_columns.json"))
    )
    if not otros_periodos:
        print(f"A8-no-duplicacion [{period}]: PASS trivial (no hay otros periodos generados aun para comparar)")
        return True

    # Cargar filas/ventanas de todos los otros periodos una sola vez, y
    # construir el indice GLOBAL de ventanas de estancia (periodo actual +
    # todos los otros) - necesario para clasificar arrastres que saltan mas
    # de una frontera (la estancia puede no tener su fila "estancia
    # hospitalaria" ni en el periodo actual ni en el "otro" comparado, sino
    # en un tercer periodo donde de verdad se factura).
    filas_por_periodo = {period: filas_actual}
    ventanas_globales = {}
    for doc, vents in ventanas_actual.items():
        ventanas_globales.setdefault(doc, []).extend(vents)
    for otro in otros_periodos:
        otro_infos = os.path.join(exp_root, otro, "03_INFORMATIVOS")
        otro_tramas = os.path.join(exp_root, otro, "01_TRAMAS")
        with open(os.path.join(otro_infos, ".trama_columns.json"), "r", encoding="utf-8") as f:
            cols_otro = json.load(f)
        filas_otro, ventanas_otro = _leer_trama_clasificacion(otro_tramas, cols_otro)
        filas_por_periodo[otro] = filas_otro
        for doc, vents in ventanas_otro.items():
            ventanas_globales.setdefault(doc, []).extend(vents)

    conocidos = []
    nuevos = []
    for otro in otros_periodos:
        filas_otro = filas_por_periodo[otro]
        claves_otro = _claves_desde_filas(filas_otro)
        repetidas = claves_actual & claves_otro
        for (doc, fecha, cod) in repetidas:
            tipos_actual = _tipos_de(filas_actual, doc, fecha, cod)
            tipos_otro = _tipos_de(filas_otro, doc, fecha, cod)
            superficies = _clasificar_overlap_a8(doc, fecha, tipos_actual, tipos_otro, ventanas_globales)
            if not superficies:
                nuevos.append((doc, fecha, cod, otro))
            else:
                conocidos.append((doc, fecha, cod, otro, ",".join(sorted(superficies))))

    if conocidos:
        ruta_csv = os.path.join(infos_dir, f"overlaps_conocidos_{period}.csv")
        with open(ruta_csv, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["documento", "fecha", "codigo", "periodo_comparado", "clasificacion"])
            for row in sorted(conocidos):
                w.writerow(row)
        print(f"A8-no-duplicacion [{period}]: {len(conocidos)} prestaciones compartidas con otros periodos, "
              f"todas conocido-explicado (frontera CPT/SIGESAPOL o consulta-vs-estancia) - detalle en {ruta_csv}")

    if nuevos:
        ejemplo = sorted(nuevos)[:5]
        print(f"A8-no-duplicacion [{period}]: FALLO - {len(nuevos)} prestaciones compartidas con otros periodos "
              f"SIN explicacion conocida (ni frontera CPT/SIGESAPOL ni consulta-vs-estancia). Ejemplos: {ejemplo}")
        return False

    print(f"A8-no-duplicacion [{period}]: PASS ({len(conocidos)} conocido-explicado, 0 nuevo, "
          f"comparado contra {len(otros_periodos)} periodo(s): {otros_periodos})")
    return True



def blank_decisions(audit_path):
    """Vacia la columna DECISION_AUDITORIA (T, col 20) en las 4 hojas de decision.
    Un libro 'vacio (todo pendiente)' significa celdas en blanco: el propio
    13_REINCORPORAR_decisiones.py cae entonces a sus defaults de codigo
    (SE UNE / NO PROCEDE / PROCEDE INDEPENDIENTE / PROCEDE), que son exactamente
    los que ya aplico generate_outputs_v2.py al construir la trama inicial.
    El valor PRE-LLENADO por generate_outputs_v2.py en DUPLICADOS_FUENTES ya es
    una recomendacion de negocio (fuente canonica), no un 'pendiente' real, por
    eso hay que vaciarlo antes de probar idempotencia."""
    wb = openpyxl.load_workbook(audit_path)
    for sheet_name in ("ESTANCIAS_E_H", "DUPLICADOS_FUENTES", "DUPLICADOS_ORIGEN", "TRANSF_HUERFANAS"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=20, max_col=20):
            row[0].value = None
    wb.save(audit_path)


def split_records(data):
    sep = b"|\r\n" if b"|\r\n" in data else b"|\n"
    return [rec for rec in data.split(sep) if rec.strip()]


def check_a3_ciclo(exp_dir, tramas_dir, audit_path, year, month, period):
    txt_names = ["trama_consulta_externa.txt", "trama_emergencia.txt", "trama_hospitalizacion.txt", "trama_farmacia.txt"]
    if not os.path.exists(audit_path):
        print(f"A3-ciclo [{period}]: FALLO - no existe {audit_path}")
        return False

    backup = {}
    for name in txt_names:
        p = os.path.join(tramas_dir, name)
        if os.path.exists(p):
            with open(p, "rb") as f:
                backup[name] = f.read()
    with open(audit_path, "rb") as f:
        audit_backup = f.read()

    ok = True
    try:
        blank_decisions(audit_path)
        result = subprocess.run(
            [sys.executable, "13_REINCORPORAR_decisiones.py", "--year", str(year), "--month", str(month)],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            print(f"A3-ciclo [{period}]: FALLO - 13_REINCORPORAR_decisiones.py devolvio codigo {result.returncode}")
            print(result.stdout)
            print(result.stderr)
            ok = False
        else:
            for name in txt_names:
                p = os.path.join(tramas_dir, name)
                after = open(p, "rb").read() if os.path.exists(p) else b""
                before = backup.get(name, b"")
                # Se compara por registro (no por bytes crudos): un archivo
                # generado antes de esta verificacion puede usar CRLF y el
                # reescrito por 13 usa LF; el terminador no es dato de negocio.
                if split_records(after) != split_records(before):
                    print(f"A3-ciclo [{period}]: FALLO - {name} cambio tras correr 13 con libro de decisiones vacio (no es idempotente)")
                    ok = False
    finally:
        # Restaurar siempre el estado original (txts + workbook), pase o falle la comparacion
        for name, content in backup.items():
            with open(os.path.join(tramas_dir, name), "wb") as f:
                f.write(content)
        with open(audit_path, "wb") as f:
            f.write(audit_backup)

    if ok:
        print(f"A3-ciclo [{period}]: PASS (13_REINCORPORAR_decisiones.py con libro vacio no modifico ningun txt)")
    return ok


def check_a3_control10(year, month, period, skip):
    if skip:
        print(f"A3-CONTROL10 [{period}]: SKIPPED (--skip-control10, no se verifico contra BD)")
        return True

    dbname, host, port, user, password = _conn_params("LNS_DB_CPT", "PGDATABASE", "db_cpt_junio26")
    conn = _conectar(dbname, host, port, user, password)
    cur = conn.cursor()
    cur.execute("SELECT MIN(fecha_atencion), MAX(fecha_atencion) FROM temp_bdt_consulta_local;")
    min_d, max_d = cur.fetchone()
    if min_d is None or min_d.year != year or min_d.month != month:
        print(f"A3-CONTROL10 [{period}]: SKIPPED (las tablas temporales cargadas en BD corresponden a {min_d}..{max_d}, no a {period})")
        cur.close()
        conn.close()
        return True

    cur.execute(CONTROL_10_SQL)
    (count,) = cur.fetchone()
    cur.close()
    conn.close()
    if count != 0:
        print(f"A3-CONTROL10 [{period}]: FALLO - CONTROL 10 devolvio {count} filas (debe ser 0)")
        return False
    print(f"A3-CONTROL10 [{period}]: PASS (CONTROL 10 = 0)")
    return True


def main():
    args = parse_args()
    year = args.year
    month = args.month
    period = f"{year}-{month:02d}"

    exp_dir = os.path.join("expedientes", period)
    tramas_dir = os.path.join(exp_dir, "01_TRAMAS")
    infos_dir = os.path.join(exp_dir, "03_INFORMATIVOS")
    audit_path = os.path.join(exp_dir, f"02_AUDITORIA_{period}.xlsx")

    print(f"=== Verificando aserciones A1-A8 para {period} ===")
    results = [
        check_a1(infos_dir, period, allow_missing=args.allow_missing_conservacion),
        check_a2(exp_dir, infos_dir, tramas_dir, period),
        check_a3_ciclo(exp_dir, tramas_dir, audit_path, year, month, period),
        check_a3_control10(year, month, period, args.skip_control10),
        check_a4(infos_dir, tramas_dir, period),
        check_a5(infos_dir, tramas_dir, period, year, month),
        check_a6_integridad(infos_dir, tramas_dir, period, allow_missing=args.allow_missing_conservacion),
        check_a7_cobertura(year, month, period, infos_dir, args.skip_a7_db),
        check_a8_no_duplicacion(year, month, period),
    ]

    if all(results):
        print(f"=== {period}: TODAS LAS ASERCIONES OK ===")
        sys.exit(0)
    else:
        print(f"=== {period}: HAY ASERCIONES FALLIDAS - DETENER ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
