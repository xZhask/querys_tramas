import os
import re
import sys
import json
import argparse
import openpyxl
import datetime

def parse_args():
    parser = argparse.ArgumentParser(description="Reincorporate Auditor decisions from Excel v2")
    parser.add_argument("--year", type=int, default=2025, help="Year of the period")
    parser.add_argument("--month", type=int, required=True, help="Month of the period (e.g. 7)")
    return parser.parse_args()

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.date) or isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        return val.strip()
    return str(val)

def load_trama_rows(filepath):
    if not os.path.exists(filepath):
        return []
    # Some source fields (diagnostic descriptions, etc.) carry stray embedded
    # \r characters. Splitting on the literal "|<newline>" record terminator
    # (as written by write_trama_file/write_trama_parts) instead of iterating
    # text-mode lines avoids misreading those embedded characters as row breaks.
    with open(filepath, 'rb') as f:
        data = f.read().decode('utf-8')
    sep = '|\r\n' if '|\r\n' in data else '|\n'
    rows = []
    for rec in data.split(sep):
        if not rec.strip():
            continue
        rows.append(rec.split('|'))
    return rows

def write_trama_parts(filepath, rows):
    # newline='' avoids re-translating stray embedded \r/\n inside field values
    # into extra \r on every rewrite (see load_trama_rows).
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        for r in rows:
            f.write("|".join(r) + "|\n")

def main():
    args = parse_args()
    year = args.year
    month = args.month
    month_str = f"{month:02d}"
    period = f"{year}-{month_str}"
    
    exp_dir = os.path.join("expedientes", period)
    tramas_dir = os.path.join(exp_dir, "01_TRAMAS")
    infos_dir = os.path.join(exp_dir, "03_INFORMATIVOS")
    audit_path = os.path.join(exp_dir, f"02_AUDITORIA_{period}.xlsx")
    log_path = os.path.join(infos_dir, "reincorporacion.log")
    
    if not os.path.exists(audit_path):
        print(f"Error: Audit workbook not found at {audit_path}")
        sys.exit(1)
        
    print(f"Reading auditor decisions for {period} from {audit_path}...")
    wb = openpyxl.load_workbook(audit_path, data_only=True)
    
    # 1. Load private logs
    with open(os.path.join(infos_dir, ".eh_groups.json"), 'r', encoding='utf-8') as f:
        eh_data = json.load(f)
        eh_groups = eh_data['groups']
        retained_emer_stays = eh_data['retained_emer_stays']
        retained_cpt_stays = eh_data['retained_cpt_stays']
        
    with open(os.path.join(infos_dir, ".retained_package_emergencia.json"), 'r', encoding='utf-8') as f:
        retained_package_eme = json.load(f)
    with open(os.path.join(infos_dir, ".retained_package_hospitalizacion.json"), 'r', encoding='utf-8') as f:
        retained_package_hosp = json.load(f)
    with open(os.path.join(infos_dir, ".retained_duplicates.json"), 'r', encoding='utf-8') as f:
        retained_duplicates = json.load(f)
        
    # Group CPT stays by id for easy lookup
    retained_cpt_stays_map = {str(r['id_prestacion_cpt']): r for r in retained_cpt_stays}
    retained_emer_stays_map = {str(r['id_atencion_emergencia']): r for r in retained_emer_stays}
    
    log_lines = []
    log_lines.append(f"LOG DE REINCORPORACION DE DECISIONES DE AUDITORIA - PERIODO {period}")
    log_lines.append(f"Fecha: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log_lines.append("="*80)
    
    # --- A. Read ESTANCIAS_E_H Decisions ---
    eh_decisions = {} # group_id -> 'SE UNE' or 'NO SE UNE'
    ws_eh = wb["ESTANCIAS_E_H"]
    for row in ws_eh.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        motivo = str(row[18] or '')
        decision = str(row[19] or 'SE UNE').strip().upper()
        m = re.search(r"GRP_\d+", motivo)
        if m:
            group_id = m.group(0)
            if group_id not in eh_decisions or decision != 'SE UNE':
                eh_decisions[group_id] = decision
                
    # --- B. Read DUPLICADOS_FUENTES Decisions ---
    dup_f_decisions = {} # (group_id, fuentes) -> 'PROCEDE' or 'NO PROCEDE'
    ws_df = wb["DUPLICADOS_FUENTES"]
    for row in ws_df.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        motivo = str(row[18] or '')
        fuente = str(row[17] or '').strip().upper()
        decision = str(row[19] or 'NO PROCEDE').strip().upper()
        m = re.search(r"DUP_F_\d+", motivo)
        if m:
            group_id = m.group(0)
            dup_f_decisions[(group_id, fuente)] = decision
            
    # --- C. Read DUPLICADOS_ORIGEN Decisions ---
    dup_o_decisions = {} # group_id -> 'CONSOLIDAR CANTIDAD' or 'PROCEDE INDEPENDIENTE'
    ws_do = wb["DUPLICADOS_ORIGEN"]
    for row in ws_do.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        motivo = str(row[18] or '')
        decision = str(row[19] or 'PROCEDE INDEPENDIENTE').strip().upper()
        m = re.search(r"DUP_O_\d+", motivo)
        if m:
            group_id = m.group(0)
            dup_o_decisions[group_id] = decision
            
    # --- D. Read TRANSF_HUERFANAS Decisions ---
    huerfanas_decisions = {} # e_id -> 'PROCEDE' or 'NO PROCEDE'
    ws_th = wb["TRANSF_HUERFANAS"]
    for row in ws_th.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        motivo = str(row[18] or '')
        decision = str(row[19] or 'PROCEDE').strip().upper()
        m = re.search(r"ID:(\d+)", motivo)
        if m:
            e_id = int(m.group(1))
            huerfanas_decisions[e_id] = decision

    print("Decisions loaded from sheets. Processing and updating tramas...")

    # Load initial tramas from files
    trama_consulta_path = os.path.join(tramas_dir, "trama_consulta_externa.txt")
    trama_emergencia_path = os.path.join(tramas_dir, "trama_emergencia.txt")
    trama_hosp_path = os.path.join(tramas_dir, "trama_hospitalizacion.txt")
    
    trama_consulta = load_trama_rows(trama_consulta_path)
    trama_emergencia = load_trama_rows(trama_emergencia_path)
    trama_hosp = load_trama_rows(trama_hosp_path)
    
    # Load exact column headers from metadata JSON
    trama_cols = {}
    cols_json_path = os.path.join(infos_dir, ".trama_columns.json")
    if os.path.exists(cols_json_path):
        with open(cols_json_path, 'r', encoding='utf-8') as f:
            trama_cols = json.load(f)
            
    def get_col_index_eme(name):
        if 'emergencia' in trama_cols and name in trama_cols['emergencia']:
            return trama_cols['emergencia'].index(name)
        # Fallback to keys of retained stays
        if retained_emer_stays:
            keys = list(retained_emer_stays[0].keys())
            if name in keys:
                return keys.index(name)
        return -1
        
    def get_col_index_hosp(name):
        if 'hospitalizacion' in trama_cols and name in trama_cols['hospitalizacion']:
            return trama_cols['hospitalizacion'].index(name)
        # Fallback to keys of retained stays
        if retained_cpt_stays:
            keys = list(retained_cpt_stays[0].keys())
            if name in keys:
                return keys.index(name)
        return -1
        
    idx_eme_id = get_col_index_eme('id_atencion_emergencia')
    idx_hosp_id = get_col_index_hosp('id_prestacion_cpt')
    idx_hosp_ing = get_col_index_hosp('sp_fecha_atencion')
    idx_hosp_alt = get_col_index_hosp('sp_fecha_alta')
    idx_hosp_days = get_col_index_hosp('sp_suma_cantidad')
    idx_hosp_val = get_col_index_hosp('sp_valorizacion_total')
    
    # --- 1. Apply ESTANCIAS_E_H Decisions ---
    log_lines.append("\n1. DECISIONES DE ESTANCIAS (Caso A):")
    log_lines.append("-" * 40)
    
    eme_stays_to_add = []
    hosp_stays_to_update = {}
    
    package_eme_procs_to_add = []
    keys_to_remove = set()
    
    for g in eh_groups:
        group_id = g['group_id']
        dec = eh_decisions.get(group_id, 'SE UNE')
        
        log_lines.append(f"Grupo {group_id} - Paciente {g['dni']} ({g['nombres_paciente']}): DECISION = {dec}")
        
        if dec == 'NO SE UNE':
            # Revert hospitalization stay to original dates
            h_id = str(g['h_id'])
            if h_id and h_id in retained_cpt_stays_map:
                r_orig = retained_cpt_stays_map[h_id]
                orig_ing = parse_date(r_orig.get('sp_fecha_atencion'))
                orig_alt = parse_date(r_orig.get('sp_fecha_alta'))
                orig_days = str(r_orig.get('sp_suma_cantidad') or 1)
                orig_val = f"{float(r_orig.get('sp_valorizacion_total') or 0.0):.2f}"
                hosp_stays_to_update[h_id] = (orig_ing, orig_alt, orig_days, orig_val)
                log_lines.append(f"  -> Revertida Estancia Hospitalizacion {h_id} a fechas originales: {orig_ing} - {orig_alt}")
            
            # Remove emergency package from hospitalisation trama
            removed_count = 0
            for row in retained_package_eme:
                if row.get('eh_group_id') == group_id:
                    key = (row.get('sp_numero_documento_paciente'), parse_date(row.get('sp_fecha_atencion')), row.get('sp_codigo_procedimiento'))
                    keys_to_remove.add(key)
                    removed_count += 1
            for row in retained_package_hosp:
                if row.get('eh_group_id') == group_id:
                    key = (row.get('sp_numero_documento_paciente'), parse_date(row.get('sp_fecha_atencion')), row.get('sp_codigo_procedimiento'))
                    keys_to_remove.add(key)
                    removed_count += 1
            log_lines.append(f"  -> Excluidos {removed_count} registros de paquete de trama Hospitalizacion.")
            
            # Restore emergency stay to emergency trama
            e_id = str(g['e_id'])
            if e_id in retained_emer_stays_map:
                eme_stays_to_add.append(retained_emer_stays_map[e_id])
                
            # Restore emergency packages to emergency trama
            restored_count = 0
            for row in retained_package_eme:
                if row.get('eh_group_id') == group_id:
                    package_eme_procs_to_add.append(row)
                    restored_count += 1
            log_lines.append(f"  -> Restablecidos {restored_count} registros a trama Emergencia.")
            
        else:
            # dec == 'SE UNE' (do nothing since it is already applied in the initial tramas!)
            pass
            
    # Apply hospitalisation stay date reverting directly in CPT stay rows
    if idx_hosp_id != -1 and idx_hosp_ing != -1 and idx_hosp_alt != -1:
        for r in trama_hosp:
            h_id = r[idx_hosp_id].strip() if len(r) > idx_hosp_id else ''
            if h_id in hosp_stays_to_update:
                new_ing, new_alt, new_days, new_val = hosp_stays_to_update[h_id]
                r[idx_hosp_ing] = new_ing
                r[idx_hosp_alt] = new_alt
                r[idx_hosp_days] = new_days
                r[idx_hosp_val] = new_val
                
    # Filter out emergency package rows (procedures/labs) from trama_hosp for NO SE UNE groups
    idx_hosp_dni = get_col_index_hosp('sp_numero_documento_paciente')
    idx_hosp_code = get_col_index_hosp('sp_codigo_procedimiento')
    idx_hosp_date = get_col_index_hosp('sp_fecha_atencion')
    
    if idx_hosp_dni != -1 and idx_hosp_code != -1 and idx_hosp_date != -1:
        filtered_trama_hosp = []
        for r in trama_hosp:
            if len(r) > max(idx_hosp_dni, idx_hosp_code, idx_hosp_date):
                dni = r[idx_hosp_dni].strip()
                date = r[idx_hosp_date].strip()
                code = r[idx_hosp_code].strip()
                if (dni, date, code) in keys_to_remove:
                    continue
            filtered_trama_hosp.append(r)
        trama_hosp = filtered_trama_hosp
                
    # --- 2. Apply DUPLICADOS_FUENTES Decisions ---
    log_lines.append("\n2. DECISIONES DE DUPLICADOS ENTRE FUENTES:")
    log_lines.append("-" * 40)
    
    dup_f_to_add = {
        'consulta': [],
        'emergencia': [],
        'hospitalizacion': []
    }
    
    for item in retained_duplicates:
        # JSON stores tuples as [row_dict, source_type]
        if isinstance(item, list) and len(item) == 2 and isinstance(item[0], dict):
            row = item[0]
            stype_hint = item[1]
        else:
            row = item
            stype_hint = None
        group_id = row['dup_group_id']
        fuente = row.get('fuentes', '')
        dec = dup_f_decisions.get((group_id, fuente), 'NO PROCEDE')
        
        base = row.get('base', '')
        stype = stype_hint or 'consulta'
        if stype_hint is None:
            if 'emergencia' in base: stype = 'emergencia'
            elif 'hospitalizacion' in base: stype = 'hospitalizacion'
        
        if dec == 'PROCEDE':
            dup_f_to_add[stype].append(row)
            log_lines.append(f"Grupo {group_id} ({fuente}) - Paciente {row.get('sp_numero_documento_paciente','')} - Proc {row.get('sp_codigo_procedimiento','')}: PROCEDE (Añadido a trama {stype})")
        else:
            log_lines.append(f"Grupo {group_id} ({fuente}) - Paciente {row.get('sp_numero_documento_paciente','')} - Proc {row.get('sp_codigo_procedimiento','')}: NO PROCEDE (Excluido)")

    # --- 3. Apply DUPLICADOS_ORIGEN Decisions ---
    log_lines.append("\n3. DECISIONES DE DUPLICADOS DE ORIGEN:")
    log_lines.append("-" * 40)
    
    # Pre-build lookup from ws_do sheet for O(1) group_id lookups
    ws_do = wb["DUPLICADOS_ORIGEN"]
    dup_o_lookup = {}  # (dni, date, code) -> group_id
    for row_xl in ws_do.iter_rows(min_row=2, values_only=True):
        if not row_xl[0]: continue
        xl_dni = str(row_xl[2] or '')
        xl_date = parse_date(row_xl[4])
        xl_code = str(row_xl[15] or '')
        motivo = str(row_xl[18] or '')
        m = re.search(r"DUP_O_\d+", motivo)
        if m:
            dup_o_lookup[(xl_dni, xl_date, xl_code)] = m.group(0)
    
    def consolidate_trama_rows(trama_rows, col_code_idx, col_dni_idx, col_date_idx, col_cant_idx, col_val_idx):
        if col_code_idx == -1 or col_dni_idx == -1 or col_date_idx == -1:
            return trama_rows

        groups = {}
        for r in trama_rows:
            if len(r) <= max(col_code_idx, col_dni_idx, col_date_idx):
                continue
            dni = r[col_dni_idx].strip()
            date = r[col_date_idx].strip()
            code = r[col_code_idx].strip()
            key = (dni, date, code)
            groups.setdefault(key, []).append(r)

        # Solo se recalculan las filas de grupos con decision CONSOLIDAR CANTIDAD;
        # el resto se deja intacto en su posicion original para que un libro de
        # decisiones vacio (todo PROCEDE INDEPENDIENTE por defecto) no reordene
        # el archivo sin necesidad.
        to_consolidate = {}
        for key, rows_list in groups.items():
            if len(rows_list) <= 1:
                continue
            group_id = dup_o_lookup.get(key)
            decision = dup_o_decisions.get(group_id, 'PROCEDE INDEPENDIENTE') if group_id else 'PROCEDE INDEPENDIENTE'
            if decision == 'CONSOLIDAR CANTIDAD':
                log_lines.append(f"Consolidando cantidades para {key[0]} - Proc {key[2]} ({len(rows_list)} duplicados): CONSOLIDAR CANTIDAD")
                first_row = list(rows_list[0])
                total_cant = sum(float(r[col_cant_idx].strip() or 0.0) for r in rows_list)
                total_val = sum(float(r[col_val_idx].strip() or 0.0) for r in rows_list)
                if total_cant.is_integer():
                    first_row[col_cant_idx] = str(int(total_cant))
                else:
                    first_row[col_cant_idx] = f"{total_cant:.2f}"
                first_row[col_val_idx] = f"{total_val:.2f}"
                to_consolidate[key] = first_row

        if not to_consolidate:
            return trama_rows

        new_rows = []
        seen_keys = set()
        for r in trama_rows:
            if len(r) <= max(col_code_idx, col_dni_idx, col_date_idx):
                new_rows.append(r)
                continue
            key = (r[col_dni_idx].strip(), r[col_date_idx].strip(), r[col_code_idx].strip())
            if key in to_consolidate:
                if key not in seen_keys:
                    new_rows.append(to_consolidate[key])
                    seen_keys.add(key)
            else:
                new_rows.append(r)
        return new_rows

    def get_col_index_consulta(name):
        if 'consulta' in trama_cols and name in trama_cols['consulta']:
            return trama_cols['consulta'].index(name)
        # Fallback
        if retained_emer_stays:
            keys = list(retained_emer_stays[0].keys())
            if name in keys:
                return keys.index(name)
        return -1
        
    idx_cons_code = get_col_index_consulta('sp_codigo_procedimiento')
    idx_cons_dni = get_col_index_consulta('sp_numero_documento_paciente')
    idx_cons_date = get_col_index_consulta('sp_fecha_atencion')
    idx_cons_cant = get_col_index_consulta('sp_suma_cantidad')
    idx_cons_val = get_col_index_consulta('sp_valorizacion_total')
    
    idx_eme_code = get_col_index_eme('sp_codigo_procedimiento')
    idx_eme_dni = get_col_index_eme('sp_numero_documento_paciente')
    idx_eme_date = get_col_index_eme('sp_fecha_atencion')
    idx_eme_cant = get_col_index_eme('sp_suma_cantidad')
    idx_eme_val = get_col_index_eme('sp_valorizacion_total')
    
    idx_hosp_code = get_col_index_hosp('sp_codigo_procedimiento')
    idx_hosp_dni = get_col_index_hosp('sp_numero_documento_paciente')
    idx_hosp_date = get_col_index_hosp('sp_fecha_atencion')
    idx_hosp_cant = get_col_index_hosp('sp_suma_cantidad')
    idx_hosp_val = get_col_index_hosp('sp_valorizacion_total')

    trama_consulta = consolidate_trama_rows(trama_consulta, idx_cons_code, idx_cons_dni, idx_cons_date, idx_cons_cant, idx_cons_val)
    trama_emergencia = consolidate_trama_rows(trama_emergencia, idx_eme_code, idx_eme_dni, idx_eme_date, idx_eme_cant, idx_eme_val)
    trama_hosp = consolidate_trama_rows(trama_hosp, idx_hosp_code, idx_hosp_dni, idx_hosp_date, idx_hosp_cant, idx_hosp_val)

    # --- 4. Apply TRANSF_HUERFANAS Decisions ---
    log_lines.append("\n4. DECISIONES DE TRANSICIONES HUERFANAS:")
    log_lines.append("-" * 40)
    
    rejected_huerfanas_ids = set()
    for e_id, dec in huerfanas_decisions.items():
        if dec == 'NO PROCEDE':
            rejected_huerfanas_ids.add(str(e_id))
            log_lines.append(f"Emergencia Huerfana ID {e_id}: NO PROCEDE (Excluida de la trama)")
            
    if idx_eme_id != -1:
        initial_count = len(trama_emergencia)
        trama_emergencia = [r for r in trama_emergencia if not (len(r) > idx_eme_id and r[idx_eme_id].strip() in rejected_huerfanas_ids)]
        log_lines.append(f"  -> Se removieron {initial_count - len(trama_emergencia)} emergencias huérfanas de trama_emergencia.txt")

    # Append all added records to respective tramas
    def format_val(v):
        if v is None: return ""
        return str(v)
        
    def append_rows_to_trama(trama_list, rows_to_add, cols_keys):
        for r in rows_to_add:
            row_list = []
            for col in cols_keys:
                val = r.get(col)
                if col == 'sp_fecha_nacimiento' and isinstance(val, (datetime.date, datetime.datetime)):
                    row_list.append(val.strftime('%d/%m/%Y'))
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    row_list.append(val.strftime('%Y-%m-%d'))
                else:
                    row_list.append(format_val(val))
            trama_list.append(row_list)

    cols_eme = list(retained_emer_stays[0].keys()) if retained_emer_stays else []
    append_rows_to_trama(trama_emergencia, eme_stays_to_add, cols_eme)
    append_rows_to_trama(trama_emergencia, package_eme_procs_to_add, cols_eme)
    
    cols_hosp = list(retained_cpt_stays[0].keys()) if retained_cpt_stays else []
    
    # Derive consultation columns from the dup rows themselves or use eme cols as fallback
    if dup_f_to_add['consulta']:
        cols_cons = list(dup_f_to_add['consulta'][0].keys())
    else:
        cols_cons = cols_eme  # Same column structure across armado scripts
    append_rows_to_trama(trama_consulta, dup_f_to_add['consulta'], cols_cons)
    append_rows_to_trama(trama_emergencia, dup_f_to_add['emergencia'], cols_eme)
    append_rows_to_trama(trama_hosp, dup_f_to_add['hospitalizacion'], cols_hosp)

    # Write updated tramas back to 01_TRAMAS/
    write_trama_parts(trama_consulta_path, trama_consulta)
    write_trama_parts(trama_emergencia_path, trama_emergencia)
    write_trama_parts(trama_hosp_path, trama_hosp)

    log_lines.append("\n" + "="*80)
    log_lines.append("FIN DEL PROCESO DE REINCORPORACION EXITOSO")
    
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(log_lines) + "\n")
        
    print(f"Reincorporation complete! Log written to {log_path}")

if __name__ == "__main__":
    main()
