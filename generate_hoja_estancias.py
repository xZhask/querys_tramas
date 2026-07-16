import os
import argparse
import datetime
import psycopg2
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def parse_args():
    parser = argparse.ArgumentParser(description="Generate auditor's Excel worksheet")
    parser.add_argument("--year", type=int, default=2025, help="Year of the period")
    parser.add_argument("--month", type=int, required=True, help="Month of the period (e.g. 7)")
    return parser.parse_args()

def connect_db():
    return psycopg2.connect("dbname=db_cpt_junio26 user=postgres password=root host=localhost")

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.date) or isinstance(val, datetime.datetime):
        # strip time
        return datetime.datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # Try formats
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return datetime.datetime(dt.year, dt.month, dt.day)
            except ValueError:
                pass
    return None

def main():
    args = parse_args()
    year = args.year
    month = args.month
    month_str = f"{month:02d}"
    period = f"{year}-{month_str}"
    
    print(f"Generating auditor's workbook for {period}...")
    
    conn = connect_db()
    cur = conn.cursor()
    
    # 1. Load CPT Hospitalizations
    # We query all columns of temp_hospitalizacion_local
    cur.execute("SELECT * FROM temp_hospitalizacion_local;")
    hosp_cols = [col[0] for col in cur.description]
    hosp_rows = cur.fetchall()
    hosp_data = [dict(zip(hosp_cols, row)) for row in hosp_rows]
    
    # 2. Load SIGESAPOL Hospitalizations (used for mismatch check)
    cur.execute("SELECT * FROM temp_hospitalizacion_sigesapol_estancia;")
    sig_hosp_cols = [col[0] for col in cur.description]
    sig_hosp_rows = cur.fetchall()
    sig_hosp_data = [dict(zip(sig_hosp_cols, row)) for row in sig_hosp_rows]
    
    # 3. Load SIGESAPOL Emergencies
    cur.execute("SELECT * FROM temp_emergencia_sigesapol_estancia;")
    emer_cols = [col[0] for col in cur.description]
    emer_rows = cur.fetchall()
    emer_data = [dict(zip(emer_cols, row)) for row in emer_rows]
    
    print(f"Loaded: {len(hosp_data)} CPT hospitalizations, {len(sig_hosp_data)} SIGESAPOL hospitalizations, {len(emer_data)} SIGESAPOL emergencies.")
    
    # Pre-parse dates to strip time and make date math easy
    for h in hosp_data:
        h['_ing_date'] = parse_date(h['sp_fecha_atencion'])
        h['_alt_date'] = parse_date(h['sp_fecha_alta'])
    for s in sig_hosp_data:
        s['_ing_date'] = parse_date(s['sp_fecha_atencion'])
        s['_alt_date'] = parse_date(s['sp_fecha_alta'])
    for e in emer_data:
        e['_ing_date'] = parse_date(e['sp_fecha_atencion'])
        e['_alt_date'] = parse_date(e['sp_fecha_alta_emergencia'])
        
    # Group CPT hospitalizations by DNI for pairing
    hosp_by_dni = {}
    for h in hosp_data:
        dni = h['sp_numero_documento_paciente']
        if dni:
            hosp_by_dni.setdefault(dni, []).append(h)
            
    # Group SIGESAPOL hospitalizations by DNI for mismatch checking
    sig_hosp_by_dni = {}
    for s in sig_hosp_data:
        dni = s['sp_numero_documento_paciente']
        if dni:
            sig_hosp_by_dni.setdefault(dni, []).append(s)
            
    # Group emergencies by DNI
    emer_by_dni = {}
    for e in emer_data:
        dni = e['sp_numero_documento_paciente']
        if dni:
            emer_by_dni.setdefault(dni, []).append(e)

    # 4. Find pairs (Caso A unions)
    # An emergency stay and a CPT hospitalization form a pair if they touch or overlap:
    # e.ingress <= h.discharge AND e.discharge >= h.ingress.
    paired_emergencies = set()
    paired_hospitalizations = set()
    
    # We will build pairs: list of tuples (e, h, min_date, max_date)
    pairs = []
    
    for h in hosp_data:
        if h['origen_reclasificacion'] == 'PERMANENCIA_EMERGENCIA_24H':
            # Caso B is not a pair with CPT hospitalization (it's a reclassified emergency)
            continue
        dni = h['sp_numero_documento_paciente']
        if not dni or dni not in emer_by_dni:
            continue
        for e in emer_by_dni[dni]:
            # Overlap/contiguity condition
            if e['_ing_date'] and e['_alt_date'] and h['_ing_date'] and h['_alt_date']:
                if e['_ing_date'] <= h['_alt_date'] and e['_alt_date'] >= h['_ing_date']:
                    # Exclude CIERRE ADMINISTRATIVO (>15 days and different month)
                    days_diff = (e['_alt_date'] - e['_ing_date']).days + 1
                    is_cierre_admin = (e['_ing_date'].strftime('%Y-%m') != e['_alt_date'].strftime('%Y-%m') and days_diff > 15)
                    if not is_cierre_admin:
                        min_date = min(e['_ing_date'], h['_ing_date'])
                        max_date = max(e['_alt_date'], h['_alt_date'])
                        pairs.append((e, h, min_date, max_date))
                        paired_emergencies.add(e['id_emergencia_sigesapol'])
                        paired_hospitalizations.add(h['id_prestacion_cpt'])

    print(f"Identified {len(pairs)} E->H pairs (unions).")

    # 5. Output blocks
    # Block 1: CPT Hospitalizations (SISTEMA=CPT, Base=HOSPITALIZACION_ESTANCIA)
    # Sorted by DNI.
    block1_rows = []
    for h in hosp_data:
        is_paired = h['id_prestacion_cpt'] in paired_hospitalizations
        estado = 'HOSP C/ EME' if is_paired else 'HOSPITALIZACION'
        
        # If paired, dates are grouped with the paired emergency
        if is_paired:
            # Find the pair dates
            my_pairs = [p for p in pairs if p[1]['id_prestacion_cpt'] == h['id_prestacion_cpt']]
            if my_pairs:
                # If multiple (rare), take the widest
                min_d = min(p[2] for p in my_pairs)
                max_d = max(p[3] for p in my_pairs)
            else:
                min_d, max_d = h['_ing_date'], h['_alt_date']
        else:
            min_d, max_d = h['_ing_date'], h['_alt_date']
            
        estancia = (max_d - min_d).days + 1 if (min_d and max_d) else h['sp_dias_estancia']
        
        row_dict = {
            'base': 'HOSPITALIZACION_ESTANCIA',
            'SISTEMA': 'CPT',
            'ESTADO': estado,
            'sp_tipo_documento_paciente': str(h['sp_tipo_documento_paciente']) if h['sp_tipo_documento_paciente'] is not None else None,
            'sp_numero_documento_paciente': h['sp_numero_documento_paciente'],
            'sp_apellido_paterno_paciente': h['sp_apellido_paterno_paciente'],
            'sp_apellido_materno_paciente': h['sp_apellido_materno_paciente'],
            'sp_nombres_paciente': h['sp_nombres_paciente'],
            'sp_fecha_nacimiento_paciente': parse_date(h['sp_fecha_nacimiento']),
            'sp_genero_paciente': str(h['sp_genero_paciente']) if h['sp_genero_paciente'] is not None else None,
            'sp_condicion_asegurado': str(h['sp_condicion_asegurado']) if h['sp_condicion_asegurado'] is not None else None,
            'sp_tipo_atencion': str(h['sp_tipo_atencion']) if h['sp_tipo_atencion'] is not None else None,
            'sp_codigo_ipress': h['sp_codigo_ipress'],
            'sp_nombre_ipress': h['sp_nombre_ipress'],
            'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION': min_d,
            'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION': max_d,
            'ESTANCIA': estancia,
            'sp_tipo_documento_responsable': str(h['sp_tipo_documento_responsable']) if h['sp_tipo_documento_responsable'] is not None else None,
            'sp_numero_documento_responsable': h['sp_numero_documento_responsable'],
            'sp_apellido_paterno_responsable': h['sp_apellido_paterno_responsable'],
            'sp_apellido_materno_responsable': h['sp_apellido_materno_responsable'],
            'sp_nombres_responsable': h['sp_nombres_responsable'],
            'sp_profesion_responsable': str(h['sp_profesion_responsable']) if h['sp_profesion_responsable'] is not None else None,
            'sp_especialidad_responsable': h['sp_especialidad_responsable'],
            'sp_circunstancia_alta': str(h['sp_circunstancia_alta']) if h['sp_circunstancia_alta'] is not None else None,
            'prioridad': None,
            'estado': None,
            'sp_upss_codigo': h['sp_upss_codigo'],
            'sp_upss_descripcion': h['sp_upss_descripcion'],
            'hospitalizacion': '1',
            'sp_tipo_dx_01': str(h['sp_tipo_dx_01']) if h['sp_tipo_dx_01'] is not None else None,
            'sp_codigo_dx_01': h['sp_codigo_dx_01'],
            'sp_descripcion_dx_01': h['sp_descripcion_dx_01'],
            'sp_tipo_dx_02': str(h['sp_tipo_dx_02']) if h['sp_tipo_dx_02'] is not None else None,
            'sp_codigo_dx_02': h['sp_codigo_dx_02'],
            'sp_descripcion_dx_02': h['sp_descripcion_dx_02'],
            'sp_tipo_dx_03': str(h['sp_tipo_dx_03']) if h['sp_tipo_dx_03'] is not None else None,
            'sp_codigo_dx_03': h['sp_codigo_dx_03'],
            'sp_descripcion_dx_03': h['sp_descripcion_dx_03'],
            'ESTADO_NOTES': None, # Mismatch note if any
            'DECISION_AUDITORIA': None
        }
        block1_rows.append(row_dict)
        
    block1_rows.sort(key=lambda x: x['sp_numero_documento_paciente'] or '')
    
    # Generate 'orden' sequential IDs for Block 1 (prefix H)
    for idx, r in enumerate(block1_rows):
        r['orden'] = f"H{idx+1:05d}"
        
    # Block 2: Paired SIGESAPOL Emergencies (SISTEMA=SIGESAPOL, Base=EMERGENCIA_ESTANCIA)
    # Also includes emergencies with CONDICION_ALTA = 3 but no hospitalization.
    # Sorted by DNI.
    block2_rows = []
    
    # We load the 13 cases (condicion_alta = 3 but no hosp)
    for e in emer_data:
        is_paired = e['id_emergencia_sigesapol'] in paired_emergencies
        is_unmatched_transfer = (e['condicion_alta'] == 3 and not is_paired and not e['excluir_tipo2'])
        
        if is_paired or is_unmatched_transfer:
            estado = 'EME C/ HOSP' if is_paired else 'EME C/ TRANS QUE NO TIENE HOSP'
            
            if is_paired:
                my_pairs = [p for p in pairs if p[0]['id_emergencia_sigesapol'] == e['id_emergencia_sigesapol']]
                if my_pairs:
                    min_d = min(p[2] for p in my_pairs)
                    max_d = max(p[3] for p in my_pairs)
                else:
                    min_d, max_d = e['_ing_date'], e['_alt_date']
            else:
                min_d, max_d = e['_ing_date'], e['_alt_date']
                
            estancia = (max_d - min_d).days + 1 if (min_d and max_d) else 1
            
            row_dict = {
                'base': 'EMERGENCIA_ESTANCIA',
                'SISTEMA': 'SIGESAPOL',
                'ESTADO': estado,
                'sp_tipo_documento_paciente': str(e['sp_tipo_documento_paciente']) if e['sp_tipo_documento_paciente'] is not None else None,
                'sp_numero_documento_paciente': e['sp_numero_documento_paciente'],
                'sp_apellido_paterno_paciente': e['sp_apellido_paterno_paciente'],
                'sp_apellido_materno_paciente': e['sp_apellido_materno_paciente'],
                'sp_nombres_paciente': e['sp_nombres_paciente'],
                'sp_fecha_nacimiento_paciente': parse_date(e['sp_fecha_nacimiento_paciente']),
                'sp_genero_paciente': str(e['sp_genero_paciente']) if e['sp_genero_paciente'] is not None else None,
                'sp_condicion_asegurado': str(e['sp_condicion_asegurado']) if e['sp_condicion_asegurado'] is not None else None,
                'sp_tipo_atencion': str(e['sp_tipo_atencion']) if e['sp_tipo_atencion'] is not None else None,
                'sp_codigo_ipress': e['sp_codigo_ipress'],
                'sp_nombre_ipress': e['sp_nombre_ipress'],
                'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION': min_d,
                'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION': max_d,
                'ESTANCIA': estancia,
                'sp_tipo_documento_responsable': str(e['sp_tipo_documento_responsable']) if e['sp_tipo_documento_responsable'] is not None else None,
                'sp_numero_documento_responsable': e['sp_numero_documento_responsable'],
                'sp_apellido_paterno_responsable': e['sp_apellido_paterno_responsable'],
                'sp_apellido_materno_responsable': e['sp_apellido_materno_responsable'],
                'sp_nombres_responsable': e['sp_nombres_responsable'],
                'sp_profesion_responsable': str(e['sp_codigo_profesion_responsable']) if e['sp_codigo_profesion_responsable'] is not None else None,
                'sp_especialidad_responsable': e['sp_codigo_especialidad'],
                'sp_circunstancia_alta': str(e['sp_circunstancia_alta_sigesapol_sp']) if e['sp_circunstancia_alta_sigesapol_sp'] is not None else None,
                'prioridad': e['prioridad'],
                'estado': e['estado'],
                'sp_upss_codigo': e['sp_upss_codigo'],
                'sp_upss_descripcion': e['sp_upss_nombre'],
                'hospitalizacion': '2',
                'sp_tipo_dx_01': str(e['sp_tipo_dx_01']) if e['sp_tipo_dx_01'] is not None else None,
                'sp_codigo_dx_01': e['sp_codigo_dx_01'],
                'sp_descripcion_dx_01': e['sp_descripcion_dx_01'],
                'sp_tipo_dx_02': str(e['sp_tipo_dx_02']) if e['sp_tipo_dx_02'] is not None else None,
                'sp_codigo_dx_02': e['sp_codigo_dx_02'],
                'sp_descripcion_dx_02': e['sp_descripcion_dx_02'],
                'sp_tipo_dx_03': str(e['sp_tipo_dx_03']) if e['sp_tipo_dx_03'] is not None else None,
                'sp_codigo_dx_03': e['sp_codigo_dx_03'],
                'sp_descripcion_dx_03': e['sp_descripcion_dx_03'],
                'ESTADO_NOTES': None,
                'DECISION_AUDITORIA': None
            }
            block2_rows.append(row_dict)
            
    block2_rows.sort(key=lambda x: x['sp_numero_documento_paciente'] or '')
    
    # Generate 'orden' sequential IDs for Block 2 (prefix E)
    for idx, r in enumerate(block2_rows):
        r['orden'] = f"E{idx+1:05d}"
        
    # Block 3: Normal SIGESAPOL Emergencies (SISTEMA=SIGESAPOL, Base=SIGESAPOL emergencia estancia)
    # Stays where excluir_tipo2 = false and not paired.
    # Sorted by DNI.
    block3_rows = []
    
    for e in emer_data:
        is_paired = e['id_emergencia_sigesapol'] in paired_emergencies
        if not is_paired and not e['excluir_tipo2']:
            # Normal emergency stays
            # Calculate note: check if patient has overlapping hospitalization (CON ESTANCIAS MONTADAS)
            note = 'EMERGENCIA'
            dni = e['sp_numero_documento_paciente']
            if dni and dni in hosp_by_dni:
                for h in hosp_by_dni[dni]:
                    # Check if overlaps
                    if e['_ing_date'] and e['_alt_date'] and h['_ing_date'] and h['_alt_date']:
                        if e['_ing_date'] <= h['_alt_date'] and e['_alt_date'] >= h['_ing_date']:
                            note = 'CON ESTANCIAS MONTADAS'
                            break
                            
            estancia = (e['_alt_date'] - e['_ing_date']).days + 1 if (e['_ing_date'] and e['_alt_date']) else 1
            
            row_dict = {
                'base': 'SIGESAPOL emergencia estancia',
                'SISTEMA': 'SIGESAPOL',
                'ESTADO': 'EMERGENCIA',
                'sp_tipo_documento_paciente': str(e['sp_tipo_documento_paciente']) if e['sp_tipo_documento_paciente'] is not None else None,
                'sp_numero_documento_paciente': e['sp_numero_documento_paciente'],
                'sp_apellido_paterno_paciente': e['sp_apellido_paterno_paciente'],
                'sp_apellido_materno_paciente': e['sp_apellido_materno_paciente'],
                'sp_nombres_paciente': e['sp_nombres_paciente'],
                'sp_fecha_nacimiento_paciente': parse_date(e['sp_fecha_nacimiento_paciente']),
                'sp_genero_paciente': str(e['sp_genero_paciente']) if e['sp_genero_paciente'] is not None else None,
                'sp_condicion_asegurado': str(e['sp_condicion_asegurado']) if e['sp_condicion_asegurado'] is not None else None,
                'sp_tipo_atencion': str(e['sp_tipo_atencion']) if e['sp_tipo_atencion'] is not None else None,
                'sp_codigo_ipress': e['sp_codigo_ipress'],
                'sp_nombre_ipress': e['sp_nombre_ipress'],
                'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION': e['_ing_date'],
                'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION': e['_alt_date'],
                'ESTANCIA': estancia,
                'sp_tipo_documento_responsable': str(e['sp_tipo_documento_responsable']) if e['sp_tipo_documento_responsable'] is not None else None,
                'sp_numero_documento_responsable': e['sp_numero_documento_responsable'],
                'sp_apellido_paterno_responsable': e['sp_apellido_paterno_responsable'],
                'sp_apellido_materno_responsable': e['sp_apellido_materno_responsable'],
                'sp_nombres_responsable': e['sp_nombres_responsable'],
                'sp_profesion_responsable': str(e['sp_codigo_profesion_responsable']) if e['sp_codigo_profesion_responsable'] is not None else None,
                'sp_especialidad_responsable': e['sp_codigo_especialidad'],
                'sp_circunstancia_alta': str(e['sp_circunstancia_alta_sigesapol_sp']) if e['sp_circunstancia_alta_sigesapol_sp'] is not None else None,
                'prioridad': e['prioridad'],
                'estado': e['estado'],
                'sp_upss_codigo': e['sp_upss_codigo'],
                'sp_upss_descripcion': e['sp_upss_nombre'],
                'hospitalizacion': '2',
                'sp_tipo_dx_01': str(e['sp_tipo_dx_01']) if e['sp_tipo_dx_01'] is not None else None,
                'sp_codigo_dx_01': e['sp_codigo_dx_01'],
                'sp_descripcion_dx_01': e['sp_descripcion_dx_01'],
                'sp_tipo_dx_02': str(e['sp_tipo_dx_02']) if e['sp_tipo_dx_02'] is not None else None,
                'sp_codigo_dx_02': e['sp_codigo_dx_02'],
                'sp_descripcion_dx_02': e['sp_descripcion_dx_02'],
                'sp_tipo_dx_03': str(e['sp_tipo_dx_03']) if e['sp_tipo_dx_03'] is not None else None,
                'sp_codigo_dx_03': e['sp_codigo_dx_03'],
                'sp_descripcion_dx_03': e['sp_descripcion_dx_03'],
                'ESTADO_NOTES': note,
                'DECISION_AUDITORIA': None
            }
            block3_rows.append(row_dict)
            
    block3_rows.sort(key=lambda x: x['sp_numero_documento_paciente'] or '')
    
    # Generate 'orden' sequential IDs for Block 3 (prefix AGRE)
    for idx, r in enumerate(block3_rows):
        r['orden'] = f"AGRE{idx+1:04d}"
        
    # Block 4: Mismatched SIGESAPOL Hospitalizations (SISTEMA=SIGESAPOL, Base=HOSPITALIZACION_ESTANCIA)
    # Stays from temp_hospitalizacion_sigesapol_estancia that overlap a CPT hospitalization but have different dates.
    # Sorted by DNI.
    block4_rows = []
    
    for h in hosp_data:
        dni = h['sp_numero_documento_paciente']
        if not dni or dni not in sig_hosp_by_dni:
            continue
        h_ing, h_alt = h['_ing_date'], h['_alt_date']
        if not h_ing or not h_alt:
            continue
        for s in sig_hosp_by_dni[dni]:
            s_ing, s_alt = s['_ing_date'], s['_alt_date']
            if not s_ing or not s_alt:
                continue
            # Check overlap
            if s_ing <= h_alt and s_alt >= h_ing:
                # Check if dates differ
                if s_ing < h_ing or s_alt > h_alt:
                    note = None
                    if s_ing < h_ing:
                        diff = (h_ing - s_ing).days
                        note = f"hos_sig_entro dias antes en sigesapol {diff}"
                    elif s_alt > h_alt:
                        diff = (s_alt - h_alt).days
                        note = f"hos_sig_se retiro despues en el sigesapol {diff}"
                        
                    estancia = (s_alt - s_ing).days + 1
                    
                    row_dict = {
                        'base': 'HOSPITALIZACION_ESTANCIA',
                        'SISTEMA': 'SIGESAPOL',
                        'ESTADO': 'HOSPITALIZACION',
                        'sp_tipo_documento_paciente': str(s['sp_tipo_documento_paciente']) if s['sp_tipo_documento_paciente'] is not None else None,
                        'sp_numero_documento_paciente': s['sp_numero_documento_paciente'],
                        'sp_apellido_paterno_paciente': s['sp_apellido_paterno_paciente'],
                        'sp_apellido_materno_paciente': s['sp_apellido_materno_paciente'],
                        'sp_nombres_paciente': s['sp_nombres_paciente'],
                        'sp_fecha_nacimiento_paciente': parse_date(s['sp_fecha_nacimiento_paciente']),
                        'sp_genero_paciente': str(s['sp_genero_paciente']) if s['sp_genero_paciente'] is not None else None,
                        'sp_condicion_asegurado': str(s['sp_condicion_asegurado']) if s['sp_condicion_asegurado'] is not None else None,
                        'sp_tipo_atencion': str(s['sp_tipo_atencion']) if s['sp_tipo_atencion'] is not None else None,
                        'sp_codigo_ipress': s['sp_codigo_ipress'],
                        'sp_nombre_ipress': s['sp_nombre_ipress'],
                        'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION': s_ing,
                        'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION': s_alt,
                        'ESTANCIA': estancia,
                        'sp_tipo_documento_responsable': str(s['sp_tipo_documento_responsable']) if s['sp_tipo_documento_responsable'] is not None else None,
                        'sp_numero_documento_responsable': s['sp_numero_documento_responsable'],
                        'sp_apellido_paterno_responsable': s['sp_apellido_paterno_responsable'],
                        'sp_apellido_materno_responsable': s['sp_apellido_materno_responsable'],
                        'sp_nombres_responsable': s['sp_nombres_responsable'],
                        'sp_profesion_responsable': str(s['sp_codigo_profesion_responsable']) if s['sp_codigo_profesion_responsable'] is not None else None,
                        'sp_especialidad_responsable': s['sp_codigo_especialidad'],
                        'sp_circunstancia_alta': str(s['sp_circunstancia_alta_sigesapol']) if s['sp_circunstancia_alta_sigesapol'] is not None else None,
                        'prioridad': None,
                        'estado': s['estado'],
                        'sp_upss_codigo': s['sp_upss_codigo'],
                        'sp_upss_descripcion': s['sp_upss_nombre'],
                        'hospitalizacion': '1',
                        'sp_tipo_dx_01': str(s['sp_tipo_dx_01']) if s['sp_tipo_dx_01'] is not None else None,
                        'sp_codigo_dx_01': s['sp_codigo_dx_01'],
                        'sp_descripcion_dx_01': s['sp_descripcion_dx_01'],
                        'sp_tipo_dx_02': str(s['sp_tipo_dx_02']) if s['sp_tipo_dx_02'] is not None else None,
                        'sp_codigo_dx_02': s['sp_codigo_dx_02'],
                        'sp_descripcion_dx_02': s['sp_descripcion_dx_02'],
                        'sp_tipo_dx_03': str(s['sp_tipo_dx_03']) if s['sp_tipo_dx_03'] is not None else None,
                        'sp_codigo_dx_03': s['sp_codigo_dx_03'],
                        'sp_descripcion_dx_03': s['sp_descripcion_dx_03'],
                        'ESTADO_NOTES': note,
                        'DECISION_AUDITORIA': None
                    }
                    block4_rows.append(row_dict)
                    
    block4_rows.sort(key=lambda x: x['sp_numero_documento_paciente'] or '')
    
    # Generate 'orden' sequential IDs for Block 4 (prefix HOSSIG)
    for idx, r in enumerate(block4_rows):
        r['orden'] = f"HOSSIG{idx+1:04d}"

    # Combine all blocks
    all_final_rows = block1_rows + block2_rows + block3_rows + block4_rows
    print(f"Final compiled sheet row count: {len(all_final_rows)}")
    print(f"  Block 1 (CPT Hosp): {len(block1_rows)}")
    print(f"  Block 2 (EME paired & unmatched transfers): {len(block2_rows)}")
    print(f"  Block 3 (EME normal): {len(block3_rows)}")
    print(f"  Block 4 (SIGESAPOL mismatch Hosp): {len(block4_rows)}")
    
    # 6. Write to Excel
    wb = openpyxl.Workbook()
    # Sheet 1: Hoja1
    ws = wb.active
    ws.title = "Hoja1"
    
    # Column headers
    headers = [
        'base', 'orden', 'SISTEMA', 'ESTADO', 'sp_tipo_documento_paciente', 'sp_numero_documento_paciente',
        'sp_apellido_paterno_paciente', 'sp_apellido_materno_paciente', 'sp_nombres_paciente',
        'sp_fecha_nacimiento_paciente', 'sp_genero_paciente', 'sp_condicion_asegurado', 'sp_tipo_atencion',
        'sp_codigo_ipress', 'sp_nombre_ipress', 'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION',
        'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION', 'ESTANCIA', 'sp_tipo_documento_responsable',
        'sp_numero_documento_responsable', 'sp_apellido_paterno_responsable', 'sp_apellido_materno_responsable',
        'sp_nombres_responsable', 'sp_profesion_responsable', 'sp_especialidad_responsable', 'sp_circunstancia_alta',
        'prioridad', 'estado', 'sp_upss_codigo', 'sp_upss_descripcion', 'hospitalizacion',
        'sp_tipo_dx_01', 'sp_codigo_dx_01', 'sp_descripcion_dx_01', 'sp_tipo_dx_02', 'sp_codigo_dx_02',
        'sp_descripcion_dx_02', 'sp_tipo_dx_03', 'sp_codigo_dx_03', 'sp_descripcion_dx_03', 'ESTADO', 'DECISION_AUDITORIA'
    ]
    ws.append(headers)
    
    # Formatting
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    for r in all_final_rows:
        row_values = []
        for h in headers:
            val = r[h] if h != 'ESTADO' else r['ESTADO']
            if h == 'ESTADO':
                # The repeated column (index 40) is for notes, index 3 is for status
                # So we put the notes column value for index 40, and status value for index 3.
                pass
            row_values.append(val)
        
        # We need to manually set the note value in column 41 (index 40)
        # and decision value in column 42 (index 41)
        row_values[40] = r['ESTADO_NOTES']
        row_values[41] = r['DECISION_AUDITORIA']
        
        ws.append(row_values)
        
    # Apply date and border formatting to rows
    date_format = 'yyyy-mm-dd'
    for r_idx in range(2, len(all_final_rows) + 2):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = border_thin
            
            # Format dates
            if headers[c_idx-1] in ['sp_fecha_nacimiento_paciente', 'sp_fecha_atencion_AGRUPADO_CON_HOSPITALIZACION', 'sp_fecha_alta_emergencia_AGRUPADO_CON_HOSPITALIZACION']:
                if isinstance(cell.value, datetime.datetime) or isinstance(cell.value, datetime.date):
                    cell.number_format = date_format
                    cell.alignment = align_center
            elif headers[c_idx-1] in ['sp_numero_documento_paciente', 'ESTANCIA', 'sp_numero_documento_responsable', 'prioridad', 'estado', 'hospitalizacion']:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format == date_format and isinstance(cell.value, datetime.datetime):
                val_str = cell.value.strftime('%Y-%m-%d')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    # Sheet 2: LEYENDA
    ws_ley = wb.create_sheet(title="LEYENDA")
    # We populate the LEYENDA sheet replicating rows 4754-4759 from the reference file
    for _ in range(4753):
        ws_ley.append([])
    ws_ley.append(['ESTADO', 'DESCRIPCION'])
    ws_ley.append(['EME C/ HOSP', 'BASE UNIDA DE EMERGENCIA Y HOSPITALIZACION, REPRESETNA EMERGENCIA'])
    ws_ley.append(['EME C/ HOSP 2', 'BASE UNIDA DE EMERGENCIA Y HOSPITALIZACION, REPRESETNA HOSPITALIZACION'])
    ws_ley.append(['EME C/ TRANS QUE NO TIENE HOSP', 'EMERGENCIA CON CIRC ALTA TRANSFERENCIA PERO NO TIENEN NINGUNA HOSPITALIZACION'])
    ws_ley.append(['HOSPITALIZACION', 'SOLO HOSPITALIZACIONES'])
    ws_ley.append(['EMERGENCIA', 'SOLO EMERGENCIAS (HAY 2 TIPOS : 1. CON ESTANCIAS MONTADAS SOBRE OTRAS PRESTACIONES 2. SOLO EMERGENCIAS QUE TIENE OBS EN LOS DIAS DE ESTANCIA'])
    
    # Save the Excel file
    output_dir = f"C:\\Users\\Intel\\Downloads\\QUERIES LNS\\ruta_querys\\expedientes\\{year}-{month_str}"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"HOJA_ESTANCIAS_AUDITORIA_{period}.xlsx")
    wb.save(output_path)
    print(f"Audit worksheet successfully saved to: {output_path}")
    
    cur.close()
    conn.close()

if __name__ == "__main__":
    main()
