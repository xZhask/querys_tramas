import os
import sys
import json
import csv
import argparse
import datetime
import psycopg2
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def parse_args():
    parser = argparse.ArgumentParser(description="Generate Rediseño de Salidas v2 outputs")
    parser.add_argument("--year", type=int, default=2025, help="Year of the period")
    parser.add_argument("--month", type=int, required=True, help="Month of the period (e.g. 7)")
    return parser.parse_args()

def connect_db():
    import os
    dbname = os.environ.get("PGDATABASE", "db_cpt_junio26")
    password = os.environ.get("PGPASSWORD", "root")
    return psycopg2.connect(f"dbname={dbname} user=postgres password={password} host=localhost")

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.date) or isinstance(val, datetime.datetime):
        return datetime.datetime(val.year, val.month, val.day)
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return datetime.datetime(dt.year, dt.month, dt.day)
            except ValueError:
                pass
    return None

def get_row_date(row):
    base = row.get('base', '')
    if 'consulta' in base:
        return parse_date(row.get('sp_fecha_atencion'))
    elif 'emergencia' in base:
        return parse_date(row.get('fecha_procedimiento') or row.get('sp_fecha_atencion'))
    elif 'hosp' in base:
        return parse_date(row.get('fecha_atencion_procedimiento') or row.get('sp_fecha_atencion'))
    return parse_date(row.get('sp_fecha_atencion'))

def get_row_doctor(row):
    base = row.get('base', '')
    if 'consulta' in base:
        return row.get('sp_numero_documento_responsable')
    elif 'emergencia' in base:
        return row.get('numero_documento_responsable_procedimiento') or row.get('sp_numero_documento_responsable')
    elif 'hosp' in base:
        return row.get('documento_responsable_cpt') or row.get('sp_numero_documento_responsable')
    return row.get('sp_numero_documento_responsable')

def get_digitador(row):
    return row.get('digitador_cpt') or row.get('digitador_laboratorio') or row.get('digitador_prestacion')

def get_unique_id(row):
    cpt_id = row.get('id_prestacion_cpt')
    if cpt_id == '': cpt_id = None
    lab_id = row.get('id_prestacion_laboratorio')
    if lab_id == '': lab_id = None
    eme_id = row.get('id_atencion_emergencia')
    if eme_id == '': eme_id = None
    return cpt_id or lab_id or eme_id or row.get('_row_idx')


def execute_sql_file(cur, filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    sql_lines = []
    for line in lines:
        l = line.strip()
        if l.startswith('--'):
            continue
        sql_lines.append(line)
    sql_content = "".join(sql_lines).strip()
    if not sql_content:
        return []
    cur.execute(sql_content)
    cols = [col[0] for col in cur.description]
    rows = cur.fetchall()
    return [dict(zip(cols, row)) for row in rows]

def format_trama_val(val):
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(val, datetime.date):
        # Check if it was originally formatted as DD/MM/YYYY
        # Actually, in the CPT database CPT dates are YYYY-MM-DD, SIGESAPOL is YYYY-MM-DD.
        # Let's keep YYYY-MM-DD for standard dates, but let's check if we should check the birth date format.
        # We will format dates as YYYY-MM-DD except if it looks like a birthday and we want to be safe.
        # Let's just output YYYY-MM-DD as standard, or DD/MM/YYYY if birth date.
        # Wait, the birth date in the old trama was DD/MM/YYYY. Let's write a simple logic:
        # If it's a date and represents sp_fecha_nacimiento, we format it as %d/%m/%Y.
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        # Saneo de exportación: algunos campos de texto libre (diagnósticos,
        # nombres) traen \r/\n embebidos desde origen. Sin esto, esos
        # caracteres quedan LITERALES dentro del valor y parten una fila
        # lógica en dos líneas físicas del .txt (62 registros de CE agosto).
        return val.replace('\r', ' ').replace('\n', ' ')
    return str(val)

def format_trama_col_val(col, val):
    # Specific format for birth dates:
    if col == 'sp_fecha_nacimiento' and isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime('%d/%m/%Y')
    return format_trama_val(val)

def write_trama_file(filepath, rows, col_keys):
    # newline='' disables universal-newline translation: some field values
    # (diagnostic descriptions, etc.) carry stray embedded \r/\n characters,
    # and text-mode translation would otherwise mangle those on every write.
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        for r in rows:
            line_parts = [format_trama_col_val(col, r.get(col)) for col in col_keys]
            # Write line with trailing pipe
            f.write("|".join(line_parts) + "|\n")

def write_trama_csv_analisis(filepath, rows, col_keys):
    # CSV de ANALISIS, no el envio oficial STIPS (ese sigue siendo
    # 01_TRAMAS/*.txt, sin cabecera, sin tocar). Este archivo es para
    # comparar contra la data de la gestion anterior: mismos valores que
    # la trama oficial, con cabecera real y columna Prestacion_ID vacia
    # (se llena manualmente despues de la revision de auditoria, fuera de
    # este pipeline). UTF-8 con BOM para que Excel muestre tildes/enes
    # correctamente sin filas fantasma.
    with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
        w = csv.writer(f)
        w.writerow(list(col_keys) + ['Prestacion_ID'])
        for r in rows:
            w.writerow([format_trama_col_val(col, r.get(col)) for col in col_keys] + [''])

def add_validation(ws, formula, cell_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error ='Valor no válido'
    dv.errorTitle = 'Selección inválida'
    dv.prompt = 'Por favor seleccione una opción de la lista'
    dv.promptTitle = 'Opciones válidas'
    ws.add_data_validation(dv)
    dv.add(cell_range)

def main():
    args = parse_args()
    year = args.year
    month = args.month
    month_str = f"{month:02d}"
    period = f"{year}-{month_str}"
    
    print(f"Executing Rediseño de Salidas v2 for period {period}...")
    
    conn = connect_db()
    cur = conn.cursor()
    
    # 1. Load CPT Hospitalization original dates for rollback
    cur.execute("SELECT id_prestacion_cpt, fecha_atencion, fecha_alta FROM temp_bdt_hospitalizacion_local;")
    orig_dates_cpt = {str(r[0]): (parse_date(r[1]), parse_date(r[2])) for r in cur.fetchall() if r[0] is not None}

    # Días/valorización ANTES de reclasificar (12_RECLASIFICAR_emergencias_24h.sql,
    # sección 1b), por row_uid - snapshot tomado antes de que el UPDATE de
    # Caso A extienda la estancia, a diferencia de orig_dates_cpt (que puede
    # traer una fila de procedimiento cualquiera con el mismo id_prestacion_cpt,
    # no necesariamente la de la estancia). Usado por 13_REINCORPORAR_decisiones.py
    # para revertir dias/valorización con exactitud en "NO SE UNE".
    cur.execute("SELECT row_uid, dias_antes, valorizacion_antes FROM temp_hospitalizacion_antes_reclasif;")
    antes_reclasif = {r[0]: (r[1], float(r[2]) if r[2] is not None else None) for r in cur.fetchall()}
    
    # Load raw stays and procedures from SQL armado scripts
    # We load them before any file outputs
    print("Loading raw data from SQL files...")
    consulta_raw = execute_sql_file(cur, "09_ARMADO_consulta_externa.sql")
    emergencia_raw = execute_sql_file(cur, "10_ARMADO_emergencia.sql")
    hospitalizacion_raw = execute_sql_file(cur, "11_ARMADO_hospitalizacion.sql")
    
    # Assign a unique row index to each record to prevent accidental multi-row exclusions
    for idx, r in enumerate(consulta_raw): r['_row_idx'] = f"C_{idx}"
    for idx, r in enumerate(emergencia_raw): r['_row_idx'] = f"E_{idx}"
    for idx, r in enumerate(hospitalizacion_raw): r['_row_idx'] = f"H_{idx}"

    # Farmacia query must run on sigesapol_junio database
    import os
    password_sig = os.environ.get("PGPASSWORD", "root")
    import os
    dbname_sig = os.environ.get("LNS_DB_SIGESAPOL")
    if not dbname_sig:
        dbname_sig = os.environ.get("PGDATABASE_SIGESAPOL", "sigesapol_junio")
    conn_sig = psycopg2.connect(f"dbname={dbname_sig} user=postgres password={password_sig} host=localhost")
    cur_sig = conn_sig.cursor()
    farmacia_raw = execute_sql_file(cur_sig, "12_SIGESAPOL_farmacia.sql")
    cur_sig.close()
    conn_sig.close()
    
    print(f"Loaded: {len(consulta_raw)} Consultation, {len(emergencia_raw)} Emergency, {len(hospitalizacion_raw)} Hospitalization, {len(farmacia_raw)} Farmacia records.")
    
    # Load E->H pairs (Caso A).
    # La unión Emergencia->Hospitalización (regla 1.4: solapa o toca) ya la
    # decidió 12_RECLASIFICAR_emergencias_24h.sql de forma determinista
    # (h.id_emergencia_unida, un solo ganador por hospitalización). Python
    # NO redefine la condición de fecha - solo consume esa unión, evitando
    # que la lista de pares del libro de auditoría pueda tener duplicados
    # internos (ver CONTEXTO_CANONICO.md §3, hallazgo "o toca").
    cur.execute("""
        SELECT
            e.id_emergencia_sigesapol,
            e.sp_numero_documento_paciente,
            e.sp_fecha_atencion AS e_ing,
            e.sp_fecha_alta_emergencia AS e_alt,
            h.id_prestacion_cpt,
            h.sp_fecha_atencion AS h_ing,
            h.sp_fecha_alta AS h_alt,
            h.row_uid,
            e.sp_apellido_paterno_paciente,
            e.sp_apellido_materno_paciente,
            e.sp_nombres_paciente,
            e.prioridad,
            e.sp_codigo_dx_01,
            e.sp_descripcion_dx_01,
            e.sp_codigo_dx_02,
            e.sp_descripcion_dx_02,
            e.sp_codigo_dx_03,
            e.sp_descripcion_dx_03
        FROM temp_hospitalizacion_local h
        JOIN temp_emergencia_sigesapol_estancia e
          ON e.id_emergencia_sigesapol = h.id_emergencia_unida
        WHERE h.origen_reclasificacion = 'UNION_EMERGENCIA_HOSP';
    """)
    pairs_rows = cur.fetchall()
    
    # Track Caso A pairs
    paired_emergencies = {} # id_emergencia_sigesapol -> CPT stay info
    paired_hospitalizations = {} # id_prestacion_cpt -> Emergency stay info
    
    eh_groups = []
    group_counter = 1
    
    for row in pairs_rows:
        e_id, dni, e_ing, e_alt, h_id, h_ing, h_alt, row_uid, pat_pat, pat_mat, pat_nom, prio, dx1, dx1_d, dx2, dx2_d, dx3, dx3_d = row
        group_id = f"GRP_{group_counter:03d}"
        group_counter += 1

        e_ing_dt = parse_date(e_ing)
        e_alt_dt = parse_date(e_alt)
        h_ing_dt = parse_date(h_ing)
        h_alt_dt = parse_date(h_alt)
        dias_antes, valorizacion_antes = antes_reclasif.get(row_uid, (None, None))
        h_ing_orig_dt = orig_dates_cpt.get(str(h_id), (h_ing_dt, h_alt_dt))[0]
        # h_alt_orig: temp_bdt_hospitalizacion_local.fecha_alta llega vacío en
        # la práctica (0/395 poblado en julio) porque esa tabla mezcla filas
        # de procedimiento con la de estancia y el diccionario se queda con
        # una fila cualquiera. dias_antes SÍ es confiable (snapshot dedicado,
        # 12_RECLASIFICAR_emergencias_24h.sql sección 1b) y da la alta real
        # sin depender de esa columna.
        h_alt_orig_dt = (
            h_ing_orig_dt + datetime.timedelta(days=dias_antes - 1)
            if h_ing_orig_dt is not None and dias_antes is not None else None
        )

        pair_info = {
            'group_id': group_id,
            'dni': dni,
            'e_id': e_id,
            'h_id': str(h_id),
            'e_ing': e_ing_dt,
            'e_alt': e_alt_dt,
            'h_ing_orig': h_ing_orig_dt,
            'h_alt_orig': h_alt_orig_dt,
            'h_dias_orig': dias_antes,
            'h_valorizacion_orig': valorizacion_antes,
            'h_ing_extended': e_ing_dt if h_ing_dt is None else min(e_ing_dt, h_ing_dt),
            'h_alt_extended': e_alt_dt if h_alt_dt is None else max(e_alt_dt, h_alt_dt),
            'prioridad': prio,
            'nombres_paciente': f"{pat_pat} {pat_mat}, {pat_nom}",
            'dx_01_codigo': dx1,
            'dx_01_descripcion': dx1_d,
            'dx_02_codigo': dx2,
            'dx_02_descripcion': dx2_d,
            'dx_03_codigo': dx3,
            'dx_03_descripcion': dx3_d
        }
        eh_groups.append(pair_info)
        paired_emergencies[e_id] = pair_info
        paired_hospitalizations[str(h_id)] = pair_info

    print(f"Found {len(eh_groups)} E->H stays solapamientos (Caso A).")

    # Group procedures and laboratories by (documento, fecha, codigo) to find duplicates
    # We unblock all procedures and labs from CE, EME and HOSP
    # Precedence: E->H package > Duplicados Fuentes > Duplicados Origen
    
    # 1. E->H package classification:
    # Any procedure/lab belonging to a paired emergency stay (same patient DNI, date within emergency stay dates)
    # is RETENIDA and marked as eh_package.
    retained_package_rows = []
    
    eh_groups_by_dni = {}
    for g in eh_groups:
        eh_groups_by_dni.setdefault(g['dni'], []).append(g)
        
    def is_in_eh_package(dni, date):
        if not dni or not date:
            return False
        for g in eh_groups_by_dni.get(dni, []):
            if g['e_ing'] <= date <= g['e_alt']:
                return g
        return None

    # Filter out E->H packages
    clean_consulta = []
    clean_emergencia = []
    clean_hospitalizacion = []
    
    # Keep track of excluded packages
    excluded_package_consulta = []
    excluded_package_emergencia = []
    excluded_package_hospitalizacion = []
    
    # Process Consultation
    for row in consulta_raw:
        # Consulta has no E->H package (it's CE, not EME/HOSP)
        clean_consulta.append(row)
        
    # Process Emergency
    for row in emergencia_raw:
        base_type = row.get('base')
        if base_type == 'estancia en emergency': # wait, spelling in sql: 'estancia en emergencia'
            # Check if this is the stay itself
            e_id = row.get('id_atencion_emergencia')
            if e_id in paired_emergencies:
                # Retained stay!
                continue
        
        # Check if it's a procedure/lab in emergency package
        dni = row.get('sp_numero_documento_paciente')
        date_at = parse_date(row.get('sp_fecha_atencion'))
        g = is_in_eh_package(dni, date_at)
        if g:
            # Belongs to emergency package! Retained!
            row['eh_group_id'] = g['group_id']
            excluded_package_emergencia.append(row)
        else:
            clean_emergencia.append(row)
            
    # Process Hospitalization
    for row in hospitalizacion_raw:
        base_type = row.get('base')
        if base_type == 'estancia hospitalaria':
            h_id = str(row.get('id_prestacion_cpt'))
            if h_id in paired_hospitalizations:
                # Paired hospitalization: we must write it to the initial tramas using its EXTENDED dates by default (SE UNE)!
                g = paired_hospitalizations[h_id]
                row_copy = dict(row)
                row_copy['sp_fecha_atencion'] = g['h_ing_extended']
                row_copy['sp_fecha_alta'] = g['h_alt_extended']
                if g['h_alt_extended'] and g['h_ing_extended']:
                    ext_days = (g['h_alt_extended'] - g['h_ing_extended']).days + 1
                else:
                    ext_days = int(row.get('sp_suma_cantidad') or 1)
                row_copy['sp_suma_cantidad'] = ext_days
                # Recalculate valuation
                rate = float(row.get('sp_valorizacion_total') or 0) / float(row.get('sp_suma_cantidad') or 1)
                row_copy['sp_valorizacion_total'] = round(ext_days * rate, 2)
                clean_hospitalizacion.append(row_copy)
                continue
        
        # Check if it's hospitalisation procedure/lab
        # Does it belong to emergency package?
        dni = row.get('sp_numero_documento_paciente')
        date_at = parse_date(row.get('sp_fecha_atencion'))
        g = is_in_eh_package(dni, date_at)
        if g and row.get('digitador_prestacion') == 'SIGESAPOL':
            # This is a procedure/lab from emergency that was copied to hospitalisation table,
            # or is part of the emergency package! Retained!
            row['eh_group_id'] = g['group_id']
            excluded_package_hospitalizacion.append(row)
        else:
            clean_hospitalizacion.append(row)

    print(f"Retained emergency stays and packages: {len(eh_groups)} stays, {len(excluded_package_emergencia) + len(excluded_package_hospitalizacion)} package records.")

    # 2. Duplicate between sources (DUPLICADOS_FUENTES) & Duplicados de Origen (DUPLICADOS_ORIGEN)
    # We group all procedures and labs from clean lists by the composite key of
    # their tipo (CONTEXTO_CANONICO.md regla #2 / checks 24-25 del piloto):
    #   Tipo 1 (consulta externa):        paciente + fecha + código + MEDICO
    #   Tipo 2 y 3 (emergencia/hosp.):     paciente + fecha + código + CANTIDAD
    # Grouping by paciente+fecha+código alone (sin el discriminador de tipo)
    # sobre-cuenta: dos atenciones legítimas del mismo código el mismo día con
    # médico o cantidad distintos NO son duplicados y deben quedar en grupos
    # separados.
    grouped_procs = {}

    # We combine procedures and labs from clean lists for duplicate checking
    # Note: we only check procedures/laboratories (base in ['procedimientos en consulta externa', 'laboratorio en consulta externa', 'procedimientos en emergencia', 'laboratorio en emergencia', 'procedimientos en hospitalización', 'laboratorio en hospitalización'])
    # Stays are NOT checked here.
    def add_to_group(row, source_type):
        base = row.get('base', '')
        if 'estancia' in base:
            return
        dni = row.get('sp_numero_documento_paciente')
        date_at = get_row_date(row)
        code = row.get('sp_codigo_procedimiento')
        if not dni or not date_at or not code:
            return

        # Normalize date to string for key
        date_str = date_at.strftime('%Y-%m-%d')
        if 'procedimientos' in base:
            discriminador = get_row_doctor(row)   # Tipo 1: +medico real
        else:
            discriminador = row.get('sp_suma_cantidad')                 # Tipo 2/3: +cantidad
            
        record_id = get_unique_id(row)
        key = (dni, date_str, code, discriminador, record_id)
        grouped_procs.setdefault(key, []).append((row, source_type))

    for r in clean_consulta: add_to_group(r, 'consulta')
    for r in clean_emergencia: add_to_group(r, 'emergencia')
    for r in clean_hospitalizacion: add_to_group(r, 'hospitalizacion')
    
    # Now classify groups
    retained_duplicates_sources = [] # list of dicts (CPT & SIGESAPOL rows)
    duplicates_origin_list = []      # list of dicts (same source rows)
    
    # Maps to track which records are excluded from clean tramas because they are retained duplicates
    excluded_duplicate_ids = set() # (source_type, id_prestacion_cpt/lab)
    
    dup_fuentes_counter = 1
    dup_origen_counter = 1
    
    for key, items in grouped_procs.items():
        # Check if duplicate between CPT and SIGESAPOL
        cpt_items = [it for it in items if get_digitador(it[0]) != 'SIGESAPOL']
        sig_items = [it for it in items if get_digitador(it[0]) == 'SIGESAPOL']
        
        if cpt_items and sig_items:
            # Duplicate between sources! RETENIDA!
            group_id = f"DUP_F_{dup_fuentes_counter:04d}"
            dup_fuentes_counter += 1
            
            # Exclude both from the initial tramas
            for row, source_type in items:
                # Mark as excluded by unique row index
                excluded_duplicate_ids.add((source_type, row['_row_idx']))
                
                # Copy row with group info for Excel
                row_copy = dict(row)
                row_copy['dup_group_id'] = group_id
                retained_duplicates_sources.append((row_copy, source_type))
                
        elif len(items) > 1:
            # Duplicate of origin! INFORMATIVA (they DO enter the tramas)
            group_id = f"DUP_O_{dup_origen_counter:04d}"
            dup_origen_counter += 1
            
            for row, source_type in items:
                row_copy = dict(row)
                row_copy['dup_group_id'] = group_id
                duplicates_origin_list.append((row_copy, source_type))

    print(f"Classified: {len(retained_duplicates_sources)//2} source duplicate pairs, {len(duplicates_origin_list)} origin duplicate records.")

    # 3. Transiciones Huérfanas (TRANSF_HUERFANAS)
    # We query the database for emergencies with condicion_alta = 3 and no hospitalization (CONTROL 12)
    # They are INFORMATIVA.
    cur.execute("""
        SELECT 
            e.id_emergencia_sigesapol,
            e.sp_numero_documento_paciente,
            e.sp_fecha_atencion,
            e.sp_fecha_alta_emergencia,
            e.sp_apellido_paterno_paciente,
            e.sp_apellido_materno_paciente,
            e.sp_nombres_paciente,
            e.prioridad,
            e.sp_codigo_dx_01,
            e.sp_descripcion_dx_01,
            e.sp_codigo_dx_02,
            e.sp_descripcion_dx_02,
            e.sp_codigo_dx_03,
            e.sp_descripcion_dx_03
        FROM temp_emergencia_sigesapol_estancia e
        WHERE e.condicion_alta = 3
          AND NOT EXISTS (
            SELECT 1 FROM temp_hospitalizacion_local h
            WHERE h.sp_numero_documento_paciente = e.sp_numero_documento_paciente
              AND e.sp_fecha_atencion::date <= h.sp_fecha_alta::date
              AND e.sp_fecha_alta_emergencia::date >= h.sp_fecha_atencion::date
          );
    """)
    huerfanas_rows = cur.fetchall()
    
    huerfanas_list = []
    for r in huerfanas_rows:
        # Build dictionary matching standard columns
        e_id, dni, e_ing, e_alt, pat_pat, pat_mat, pat_nom, prio, dx1, dx1_d, dx2, dx2_d, dx3, dx3_d = r
        prio_code = '99281'
        if prio == 1: prio_code = '99285'
        elif prio == 2: prio_code = '99284'
        elif prio == 3: prio_code = '99282'
        elif prio == 4: prio_code = '99281'
        
        huerfanas_list.append({
            'periodo': period,
            'tipo_atencion': '2',
            'documento_paciente': dni,
            'nombres_paciente': f"{pat_pat} {pat_mat}, {pat_nom}",
            'fecha_ingreso': parse_date(e_ing),
            'fecha_alta': parse_date(e_alt),
            'dx_01_tipo': '2',
            'dx_01_codigo': dx1,
            'dx_01_descripcion': dx1_d,
            'dx_02_tipo': '2' if dx2 else None,
            'dx_02_codigo': dx2,
            'dx_02_descripcion': dx2_d,
            'dx_03_tipo': '2' if dx3 else None,
            'dx_03_codigo': dx3,
            'dx_03_descripcion': dx3_d,
            'codigo_procedimiento': prio_code,
            'descripcion_procedimiento': f"Consulta en emergencia prioridad {prio}",
            'fuentes': 'SIGESAPOL',
            'MOTIVO': 'TRANSFERENCIA SIN HOSPITALIZACION',
            'DECISION_AUDITORIA': None,
            'OBSERVACION_AUDITORIA': None,
            'e_id': e_id
        })

    print(f"Found {len(huerfanas_list)} transferencias huérfanas (INFORMATIVA).")

    # 4. Filter clean lists by removing retained duplicates
    final_consulta = [r for r in clean_consulta if not (( 'consulta', r['_row_idx'] ) in excluded_duplicate_ids)]
    final_emergencia = [r for r in clean_emergencia if not (( 'emergencia', r['_row_idx'] ) in excluded_duplicate_ids)]
    final_hospitalizacion = [r for r in clean_hospitalizacion if not (( 'hospitalizacion', r['_row_idx'] ) in excluded_duplicate_ids)]
    final_farmacia = list(farmacia_raw) # Farmacia is always clean
    
    # Append emergency package rows directly to final_hospitalizacion (SE UNE default)
    for r in excluded_package_emergencia:
        r_copy = dict(r)
        final_hospitalizacion.append(r_copy)
    for r in excluded_package_hospitalizacion:
        r_copy = dict(r)
        final_hospitalizacion.append(r_copy)

    print(f"Final clean lines to write: CE={len(final_consulta)}, EME={len(final_emergencia)}, HOSP={len(final_hospitalizacion)}, FARM={len(final_farmacia)}")

    # A2 (PAQUETE COMPLETO): ningun procedimiento/laboratorio de una estancia RETENIDA
    # (paquete Caso A movido a hospitalizacion) debe quedar huerfano en trama_emergencia.
    def row_key(r):
        return (r.get('sp_numero_documento_paciente'), parse_date(r.get('sp_fecha_atencion')), r.get('sp_codigo_procedimiento'))
    retained_package_keys = {row_key(r) for r in excluded_package_emergencia} | {row_key(r) for r in excluded_package_hospitalizacion}
    leaked_in_emergencia = [r for r in final_emergencia if row_key(r) in retained_package_keys]
    if leaked_in_emergencia:
        print(f"A2 FALLIDA: {len(leaked_in_emergencia)} procedimientos/laboratorio de estancias RETENIDAS (Caso A) quedaron en trama_emergencia en vez de moverse a trama_hospitalizacion.")
        sys.exit(1)
    print("A2 (paquete completo): OK, cero fugas de paquete retenido en trama_emergencia.")

    # A1 (CONSERVACION): LIMPIA + RETENIDA + INFORMATIVA = total extraido, por tipo de trama.
    retenida_por_tipo = {'consulta': 0, 'emergencia': 0, 'hospitalizacion': 0}
    for _, stype in retained_duplicates_sources:
        retenida_por_tipo[stype] = retenida_por_tipo.get(stype, 0) + 1
    # Caso A (regla 8 del ancla): los pares E->H propuestos por el pipeline
    # son RETENIDA, no facturacion aplicada, aunque por defecto ("SE UNE") el
    # pipeline ya los escriba en trama_hospitalizacion con fecha extendida a
    # la espera de que Auditoria Medica confirme PROCEDE/NO PROCEDE. Sin esto
    # se contaban como LIMPIA. Cuentan: las propias estancias unidas
    # (eh_groups) mas el paquete de procedimientos/laboratorio movido.
    retenida_por_tipo['hospitalizacion'] += len(eh_groups)
    retenida_por_tipo['hospitalizacion'] += len(excluded_package_emergencia) + len(excluded_package_hospitalizacion)
    informativa_por_tipo = {'consulta': 0, 'emergencia': 0, 'hospitalizacion': 0}
    for _, stype in duplicates_origin_list:
        informativa_por_tipo[stype] = informativa_por_tipo.get(stype, 0) + 1
    informativa_por_tipo['emergencia'] += len(huerfanas_list)

    total_extraido = {
        'consulta': len(consulta_raw),
        'emergencia': len(emergencia_raw) - len(excluded_package_emergencia),
        'hospitalizacion': len(hospitalizacion_raw) + len(excluded_package_emergencia),
        'farmacia': len(farmacia_raw)
    }
    final_counts = {
        'consulta': len(final_consulta),
        'emergencia': len(final_emergencia),
        'hospitalizacion': len(final_hospitalizacion),
        'farmacia': len(final_farmacia)
    }

    conservacion = {}
    conservacion_failed = []
    for tipo in ('consulta', 'emergencia', 'hospitalizacion', 'farmacia'):
        retenida = retenida_por_tipo.get(tipo, 0)
        informativa = informativa_por_tipo.get(tipo, 0)
        limpia = final_counts[tipo] - retenida - informativa
        total = total_extraido[tipo]
        residuo = total - (limpia + retenida + informativa)
        conservacion[tipo] = {
            'limpia': limpia,
            'retenida': retenida,
            'informativa': informativa,
            'total_extraido': total,
            'residuo': residuo
        }
        if residuo != 0:
            conservacion_failed.append(tipo)

    if conservacion_failed:
        print(f"A1 FALLIDA: residuo distinto de cero en tramas: {conservacion_failed}")
        print(json.dumps(conservacion, indent=2))
        sys.exit(1)
    print("A1 (conservacion): OK, residuo = 0 en las 4 tramas.")

    # 5. Export clean tramas to 01_TRAMAS/
    exp_dir = os.path.join("expedientes", period)
    tramas_dir = os.path.join(exp_dir, "01_TRAMAS")
    infos_dir = os.path.join(exp_dir, "03_INFORMATIVOS")
    analisis_dir = os.path.join(exp_dir, "04_ANALISIS")
    os.makedirs(tramas_dir, exist_ok=True)
    os.makedirs(infos_dir, exist_ok=True)
    os.makedirs(analisis_dir, exist_ok=True)
    
    # We get column keys from headers in SQL files
    cur.execute("SELECT * FROM temp_bdt_consulta_local LIMIT 0;")
    # Wait! The columns list must match the select statement in 09_ARMADO_consulta_externa.sql
    # We can just extract them from our execute_sql_file description or keys!
    # Let's see, what are the keys of rows?
    # In Python, dict keys preserve the order of the columns in select statement in newer Python versions!
    # So we can just use the keys of the first row!
    cols_consulta = [k for k in list(consulta_raw[0].keys()) if k not in ('_row_idx', 'dup_group_id', 'eh_group_id')] if consulta_raw else []
    cols_emergencia = [k for k in list(emergencia_raw[0].keys()) if k not in ('_row_idx', 'dup_group_id', 'eh_group_id')] if emergencia_raw else []
    cols_hospitalizacion = [k for k in list(hospitalizacion_raw[0].keys()) if k not in ('_row_idx', 'dup_group_id', 'eh_group_id')] if hospitalizacion_raw else []
    cols_farmacia = [k for k in list(farmacia_raw[0].keys()) if k not in ('_row_idx', 'dup_group_id', 'eh_group_id')] if farmacia_raw else []
    
    with open(os.path.join(infos_dir, ".trama_columns.json"), 'w', encoding='utf-8') as f:
        json.dump({
            'consulta': cols_consulta,
            'emergencia': cols_emergencia,
            'hospitalizacion': cols_hospitalizacion,
            'farmacia': cols_farmacia
        }, f, default=str)
        
    write_trama_file(os.path.join(tramas_dir, "trama_consulta_externa.txt"), final_consulta, cols_consulta)
    write_trama_file(os.path.join(tramas_dir, "trama_emergencia.txt"), final_emergencia, cols_emergencia)
    write_trama_file(os.path.join(tramas_dir, "trama_hospitalizacion.txt"), final_hospitalizacion, cols_hospitalizacion)
    write_trama_file(os.path.join(tramas_dir, "trama_farmacia.txt"), final_farmacia, cols_farmacia)

    # CSV de analisis (cabecera real + Prestacion_ID vacia, UTF-8 con BOM) -
    # NO es el envio oficial STIPS, es para comparar contra la data de la
    # gestion anterior. Ver diccionario_tramas.md.
    write_trama_csv_analisis(os.path.join(analisis_dir, "trama_consulta_externa_analisis.csv"), final_consulta, cols_consulta)
    write_trama_csv_analisis(os.path.join(analisis_dir, "trama_emergencia_analisis.csv"), final_emergencia, cols_emergencia)
    write_trama_csv_analisis(os.path.join(analisis_dir, "trama_hospitalizacion_analisis.csv"), final_hospitalizacion, cols_hospitalizacion)
    write_trama_csv_analisis(os.path.join(analisis_dir, "trama_farmacia_analisis.csv"), final_farmacia, cols_farmacia)
    
    print("Clean tramas successfully written.")

    # 6. Save private retained package and duplicate JSON files in 03_INFORMATIVOS/
    # This allows script 13 to easily read and restore them!
    with open(os.path.join(infos_dir, ".retained_package_emergencia.json"), 'w', encoding='utf-8') as f:
        json.dump(excluded_package_emergencia, f, default=str)
    with open(os.path.join(infos_dir, ".retained_package_hospitalizacion.json"), 'w', encoding='utf-8') as f:
        json.dump(excluded_package_hospitalizacion, f, default=str)
    with open(os.path.join(infos_dir, ".retained_duplicates.json"), 'w', encoding='utf-8') as f:
        json.dump(retained_duplicates_sources, f, default=str)
    with open(os.path.join(infos_dir, ".eh_groups.json"), 'w', encoding='utf-8') as f:
        # We save stays info
        # Let's save the original stay rows for Emergency and CPT hospitalization too!
        # First find emergency stay rows. NO se pueden tomar de emergencia_raw:
        # la rama "estancia en emergencia" de 10_ARMADO_emergencia.sql filtra
        # WHERE e.excluir_tipo2 = false, y TODA emergencia unida en Caso A
        # tiene excluir_tipo2 = true (paso 8) - por construcción esa fila
        # nunca se extrae, y "NO SE UNE" en 13_REINCORPORAR_decisiones.py no
        # tenía nada que restaurar (0 filas para las 395 uniones de julio,
        # verificado). Se vuelve a consultar directo, sin el filtro, para las
        # emergencias efectivamente unidas (paired_emergencies) - mismas
        # columnas/orden que la rama original de 10_ARMADO_emergencia.sql.
        retained_emer_stays = []
        if paired_emergencies:
            cur.execute("""
                SELECT
                    'estancia en emergencia'::text as base,
                    prioridad,
                    sp_tipo_documento_paciente, sp_numero_documento_paciente,
                    sp_apellido_paterno_paciente, sp_apellido_materno_paciente, sp_nombres_paciente,
                    sp_fecha_nacimiento_paciente AS sp_fecha_nacimiento, sp_genero_paciente, sp_condicion_asegurado, sp_tipo_atencion,
                    sp_codigo_ipress, sp_nombre_ipress, e.sp_fecha_atencion, e.sp_fecha_alta_emergencia AS sp_fecha_alta,
                    sp_tipo_documento_responsable, sp_numero_documento_responsable,
                    sp_apellido_paterno_responsable, sp_apellido_materno_responsable, sp_nombres_responsable,
                    sp_codigo_profesion_responsable AS sp_profesion_responsable, sp_codigo_especialidad AS sp_especialidad_responsable,
                    sp_circunstancia_alta_sigesapol_sp::int AS sp_circunstancia_alta,
                    sp_upss_codigo,
                    regexp_replace(sp_upss_nombre, '\\r|\\n', '', 'g') as sp_upss_descripcion,
                    '2' AS hospitalizacion,
                    sp_tipo_dx_01, sp_codigo_dx_01, sp_descripcion_dx_01,
                    sp_tipo_dx_02, sp_codigo_dx_02, sp_descripcion_dx_02,
                    sp_tipo_dx_03, sp_codigo_dx_03, sp_descripcion_dx_03,
                    'SIGESAPOL' AS digitador_prestacion,
                    e.sp_fecha_atencion::date AS fecha_registro_prestacion,
                    NULL::time AS hora_registro_prestacion,
                    id_emergencia_sigesapol AS id_atencion_emergencia,
                    CASE
                        WHEN e.prioridad = 1 THEN '99285'
                        WHEN e.prioridad = 2 THEN '99284'
                        WHEN e.prioridad = 3 THEN '99282'
                        WHEN e.prioridad = 4 THEN '99281'
                        ELSE '99281'
                    END as sp_codigo_procedimiento,
                    CASE
                        WHEN e.prioridad = 1 THEN 'Consulta en emergencia para evaluación y manejo de un paciente (Prioridad I)'
                        WHEN e.prioridad = 2 THEN 'Consulta en emergencia para evaluación y manejo de un paciente (Prioridad II)'
                        WHEN e.prioridad = 3 THEN 'Consulta en emergencia para evaluación y manejo de un paciente (Prioridad III)'
                        WHEN e.prioridad = 4 THEN 'Consulta en emergencia para evaluación y manejo de un paciente (Prioridad IV)'
                        ELSE 'Consulta en emergencia'
                    END as sp_descripcion_procedimiento,
                    1 AS sp_suma_cantidad,
                    COALESCE((SELECT nivel_3 FROM cpt WHERE cod_cpt = (
                        CASE
                            WHEN e.prioridad = 1 THEN '99285'
                            WHEN e.prioridad = 2 THEN '99284'
                            WHEN e.prioridad = 3 THEN '99282'
                            WHEN e.prioridad = 4 THEN '99281'
                            ELSE '99281'
                        END
                    ) LIMIT 1), 15.31) as sp_valorizacion_total,
                    sp_numero_documento_responsable as documento_responsable_cpt,
                    concat(sp_apellido_paterno_responsable,' ',sp_apellido_materno_responsable,', ',sp_nombres_responsable) as nombre_responsable_cpt,
                    sp_upss_codigo as upss_codigo_cpt,
                    regexp_replace(sp_upss_nombre, '\\r|\\n', '', 'g') as upss_descripcion_cpt,
                    e.sp_fecha_atencion as fecha_procedimiento,
                    sp_upss_codigo as upss_codigo_procedimiento,
                    regexp_replace(sp_upss_nombre, '\\r|\\n', '', 'g') as upss_descripcion_procedimiento,
                    sp_numero_documento_responsable as numero_documento_responsable_procedimiento,
                    sp_apellido_paterno_responsable as apellido_paterno_responsable_procedimiento,
                    sp_apellido_materno_responsable as apellido_materno_responsable_procedimiento,
                    sp_nombres_responsable as nombres_responsable_procedimiento,
                    'SIGESAPOL' as digitador_cpt, e.sp_fecha_atencion::date as fecha_registro_cpt, NULL::time as hora_registro_cpt,
                    ''::text as id_prestacion_cpt,
                    ''::text as id_prestacion_laboratorio
                FROM temp_emergencia_sigesapol_estancia e
                WHERE e.id_emergencia_sigesapol = ANY(%s);
            """, (list(paired_emergencies.keys()),))
            cols_retained_eme = [col[0] for col in cur.description]
            retained_emer_stays = [dict(zip(cols_retained_eme, row)) for row in cur.fetchall()]
                    
        # Find original CPT stay rows:
        retained_cpt_stays = []
        for r in hospitalizacion_raw:
            if r.get('base') == 'estancia hospitalaria':
                h_id = str(r.get('id_prestacion_cpt'))
                if h_id in paired_hospitalizations:
                    retained_cpt_stays.append(r)
                    
        json.dump({
            'groups': eh_groups,
            'retained_emer_stays': retained_emer_stays,
            'retained_cpt_stays': retained_cpt_stays
        }, f, default=str)
        
    print("Private logs for reincorporation successfully written.")

    # 7. Generate Excel 02_AUDITORIA_<AAAA-MM>.xlsx
    audit_path = os.path.join(exp_dir, f"02_AUDITORIA_{period}.xlsx")
    wb = openpyxl.Workbook()
    
    headers = [
        'periodo', 'tipo_atencion', 'documento_paciente', 'nombres_paciente',
        'fecha_ingreso', 'fecha_alta', 'dx_01_tipo', 'dx_01_codigo', 'dx_01_descripcion',
        'dx_02_tipo', 'dx_02_codigo', 'dx_02_descripcion', 'dx_03_tipo', 'dx_03_codigo', 'dx_03_descripcion',
        'codigo_procedimiento', 'descripcion_procedimiento', 'fuentes', 'MOTIVO', 'DECISION_AUDITORIA', 'OBSERVACION_AUDITORIA'
    ]
    
    # Helper to write sheet
    def create_audit_sheet(ws, title, rows_data, decision_formula=None):
        ws.title = title
        ws.append(headers)
        
        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        border_thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        font_cell = Font(name="Calibri", size=10)
        
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        date_format = 'yyyy-mm-dd'
        col_widths = [0] * len(headers)
        
        for r_idx, r in enumerate(rows_data, 2):
            for c_idx, h in enumerate(headers, 1):
                val = r.get(h)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_cell
                cell.border = border_thin
                
                # Alignments and Formats
                if h in ['fecha_ingreso', 'fecha_alta']:
                    if isinstance(val, (datetime.datetime, datetime.date)):
                        cell.number_format = date_format
                    cell.alignment = align_center
                elif h in ['periodo', 'tipo_atencion', 'documento_paciente', 'dx_01_codigo', 'dx_02_codigo', 'dx_03_codigo', 'codigo_procedimiento', 'fuentes']:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
                val_str = str(val or '')
                if len(val_str) > col_widths[c_idx-1]:
                    col_widths[c_idx-1] = len(val_str)
                    
        # Apply dropdown list validation on column T (20) once for the whole column
        if decision_formula and rows_data:
            max_row = len(rows_data) + 1
            dv = DataValidation(type="list", formula1=decision_formula, allow_blank=True)
            dv.error ='Valor no válido'
            dv.errorTitle = 'Selección inválida'
            dv.prompt = 'Por favor seleccione una opción de la lista'
            dv.promptTitle = 'Opciones válidas'
            ws.add_data_validation(dv)
            dv.add(f"T2:T{max_row}")
            
        # Auto adjust column widths
        for c_idx, width in enumerate(col_widths, 1):
            col_letter = openpyxl.utils.get_column_letter(c_idx)
            ws.column_dimensions[col_letter].width = max(width + 3, 10)
            
    # Sheet 1: ESTANCIAS_E_H
    # Convert E->H pairs to sheet rows
    eh_sheet_rows = []
    for g in eh_groups:
        # Emergency stay row
        eh_sheet_rows.append({
            'periodo': period,
            'tipo_atencion': '2',
            'documento_paciente': g['dni'],
            'nombres_paciente': g['nombres_paciente'],
            'fecha_ingreso': g['e_ing'],
            'fecha_alta': g['e_alt'],
            'dx_01_tipo': '2',
            'dx_01_codigo': g['dx_01_codigo'],
            'dx_01_descripcion': g['dx_01_descripcion'],
            'dx_02_tipo': '2' if g['dx_02_codigo'] else None,
            'dx_02_codigo': g['dx_02_codigo'],
            'dx_02_descripcion': g['dx_02_descripcion'],
            'dx_03_tipo': '2' if g['dx_03_codigo'] else None,
            'dx_03_codigo': g['dx_03_codigo'],
            'dx_03_descripcion': g['dx_03_descripcion'],
            'codigo_procedimiento': '99281', # Mapped CPMS code of stay
            'descripcion_procedimiento': 'Consulta de emergencia',
            'fuentes': 'SIGESAPOL',
            'MOTIVO': f"SOLAPAMIENTO E-H Group {g['group_id']}",
            'DECISION_AUDITORIA': 'SE UNE', # Default decision
            'OBSERVACION_AUDITORIA': ''
        })
        # CPT hospitalisation stay row
        eh_sheet_rows.append({
            'periodo': period,
            'tipo_atencion': '3',
            'documento_paciente': g['dni'],
            'nombres_paciente': g['nombres_paciente'],
            'fecha_ingreso': g['h_ing_orig'],
            'fecha_alta': g['h_alt_orig'],
            'dx_01_tipo': '2',
            'dx_01_codigo': g['dx_01_codigo'],
            'dx_01_descripcion': g['dx_01_descripcion'],
            'dx_02_tipo': '2' if g['dx_02_codigo'] else None,
            'dx_02_codigo': g['dx_02_codigo'],
            'dx_02_descripcion': g['dx_02_descripcion'],
            'dx_03_tipo': '2' if g['dx_03_codigo'] else None,
            'dx_03_codigo': g['dx_03_codigo'],
            'dx_03_descripcion': g['dx_03_descripcion'],
            'codigo_procedimiento': '99231', # Hospital stay CPT
            'descripcion_procedimiento': 'Estancia hospitalaria CPT',
            'fuentes': 'CPT',
            'MOTIVO': f"SOLAPAMIENTO E-H Group {g['group_id']}",
            'DECISION_AUDITORIA': 'SE UNE', # Default decision
            'OBSERVACION_AUDITORIA': ''
        })
        
    ws_eh = wb.active
    create_audit_sheet(ws_eh, "ESTANCIAS_E_H", eh_sheet_rows, '"SE UNE,NO SE UNE"')
    
    # Sheet 2: DUPLICADOS_FUENTES
    # Convert duplicate procedures/labs to sheet rows
    # Sorted by group_id and source to place CPT and SIGESAPOL contiguously
    dup_f_rows = []
    # Sort retained duplicates by group id
    retained_duplicates_sources.sort(key=lambda x: x[0]['dup_group_id'])
    for r, stype in retained_duplicates_sources:
        # Determine if CPT or SIGESAPOL
        is_sig = get_digitador(r) == 'SIGESAPOL'
        default_decision = 'NO PROCEDE' if is_sig else 'PROCEDE' # default to keep CPT for pilot or vice-versa
        # Let's check canonico:
        # If month >= 10: SIGESAPOL is canonical, so default is PROCEDE for SIGESAPOL, NO PROCEDE for CPT
        if month >= 10:
            default_decision = 'PROCEDE' if is_sig else 'NO PROCEDE'
            
        dup_f_rows.append({
            'periodo': period,
            'tipo_atencion': str(r.get('sp_tipo_atencion')),
            'documento_paciente': r.get('sp_numero_documento_paciente'),
            'nombres_paciente': f"{r.get('sp_apellido_paterno_paciente')} {r.get('sp_apellido_materno_paciente')}, {r.get('sp_nombres_paciente')}",
            'fecha_ingreso': parse_date(r.get('sp_fecha_atencion')),
            'fecha_alta': parse_date(r.get('sp_fecha_alta')),
            'dx_01_tipo': r.get('sp_tipo_dx_01') or r.get('sp_tipo_diagnostico'),
            'dx_01_codigo': r.get('sp_codigo_dx_01') or r.get('sp_codigo_diagnostico'),
            'dx_01_descripcion': r.get('sp_descripcion_dx_01') or r.get('sp_descripcion_diagnostico'),
            'dx_02_tipo': r.get('sp_tipo_dx_02'),
            'dx_02_codigo': r.get('sp_codigo_dx_02'),
            'dx_02_descripcion': r.get('sp_descripcion_dx_02'),
            'dx_03_tipo': r.get('sp_tipo_dx_03'),
            'dx_03_codigo': r.get('sp_codigo_dx_03'),
            'dx_03_descripcion': r.get('sp_descripcion_dx_03'),
            'codigo_procedimiento': r.get('sp_codigo_procedimiento'),
            'descripcion_procedimiento': r.get('sp_descripcion_procedimiento'),
            'fuentes': 'SIGESAPOL' if is_sig else 'CPT',
            'MOTIVO': f"DUPLICADO ENTRE FUENTES Group {r['dup_group_id']}",
            'DECISION_AUDITORIA': default_decision,
            'OBSERVACION_AUDITORIA': ''
        })
        
    ws_df = wb.create_sheet()
    create_audit_sheet(ws_df, "DUPLICADOS_FUENTES", dup_f_rows, '"PROCEDE,NO PROCEDE"')
    
    # Sheet 3: DUPLICADOS_ORIGEN
    dup_o_rows = []
    duplicates_origin_list.sort(key=lambda x: x[0]['dup_group_id'])
    for r, stype in duplicates_origin_list:
        is_sig = get_digitador(r) == 'SIGESAPOL'
        dup_o_rows.append({
            'periodo': period,
            'tipo_atencion': str(r.get('sp_tipo_atencion')),
            'documento_paciente': r.get('sp_numero_documento_paciente'),
            'nombres_paciente': f"{r.get('sp_apellido_paterno_paciente')} {r.get('sp_apellido_materno_paciente')}, {r.get('sp_nombres_paciente')}",
            'fecha_ingreso': parse_date(r.get('sp_fecha_atencion')),
            'fecha_alta': parse_date(r.get('sp_fecha_alta')),
            'dx_01_tipo': r.get('sp_tipo_dx_01') or r.get('sp_tipo_diagnostico'),
            'dx_01_codigo': r.get('sp_codigo_dx_01') or r.get('sp_codigo_diagnostico'),
            'dx_01_descripcion': r.get('sp_descripcion_dx_01') or r.get('sp_descripcion_diagnostico'),
            'dx_02_tipo': r.get('sp_tipo_dx_02'),
            'dx_02_codigo': r.get('sp_codigo_dx_02'),
            'dx_02_descripcion': r.get('sp_descripcion_dx_02'),
            'dx_03_tipo': r.get('sp_tipo_dx_03'),
            'dx_03_codigo': r.get('sp_codigo_dx_03'),
            'dx_03_descripcion': r.get('sp_descripcion_dx_03'),
            'codigo_procedimiento': r.get('sp_codigo_procedimiento'),
            'descripcion_procedimiento': r.get('sp_descripcion_procedimiento'),
            'fuentes': 'SIGESAPOL' if is_sig else 'CPT',
            'MOTIVO': f"DUPLICADO DE ORIGEN Group {r['dup_group_id']}",
            'DECISION_AUDITORIA': 'PROCEDE INDEPENDIENTE', # Default
            'OBSERVACION_AUDITORIA': ''
        })
    ws_do = wb.create_sheet()
    create_audit_sheet(ws_do, "DUPLICADOS_ORIGEN", dup_o_rows, '"CONSOLIDAR CANTIDAD,PROCEDE INDEPENDIENTE"')
    
    # Sheet 4: TRANSF_HUERFANAS
    ws_th = wb.create_sheet()
    create_audit_sheet(ws_th, "TRANSF_HUERFANAS", huerfanas_list, '"PROCEDE,NO PROCEDE"')
    
    # Sheet 5: LEYENDA
    ws_ley = wb.create_sheet(title="LEYENDA")
    ws_ley.append(['ESTADO', 'DECISION_AUDITORIA', 'DESCRIPCION'])
    ws_ley.append(['ESTANCIAS_E_H', 'SE UNE', 'Se unifica la estancia de emergencia con la de hospitalización (se amplía fecha de ingreso hosp. y se reincorpora el paquete de cobro de emergencia en hospitalización).'])
    ws_ley.append(['ESTANCIAS_E_H', 'NO SE UNE', 'Se rompe la unión. La emergencia vuelve a la trama Tipo 2 con código CPMS por prioridad y cantidad 1. Su paquete de cobro se factura en emergencia.'])
    ws_ley.append(['DUPLICADOS_FUENTES', 'PROCEDE', 'Se autoriza la inclusión de este registro en la trama correspondiente.'])
    ws_ley.append(['DUPLICADOS_FUENTES', 'NO PROCEDE', 'Se descarta el registro por completo de la facturación.'])
    ws_ley.append(['DUPLICADOS_ORIGEN', 'CONSOLIDAR CANTIDAD', 'Se consolidan las cantidades de los registros duplicados de origen en un solo registro de facturación sumando cantidades.'])
    ws_ley.append(['DUPLICADOS_ORIGEN', 'PROCEDE INDEPENDIENTE', 'Se envían los registros por separado a las tramas finales tal como están.'])
    ws_ley.append(['TRANSF_HUERFANAS', 'PROCEDE', 'Se aprueba el envío de la emergencia a facturación Tipo 2.'])
    ws_ley.append(['TRANSF_HUERFANAS', 'NO PROCEDE', 'Se elimina la emergencia de la facturación.'])
    
    wb.save(audit_path)
    print(f"Workbook successfully written to: {audit_path}")
    
    # Also write a standard controles_integridad.txt file in 03_INFORMATIVOS/
    # We can just write a short status summary
    with open(os.path.join(infos_dir, "controles_integridad.txt"), "w", encoding="utf-8") as f:
        f.write(f"PROCESAMIENTO DE REGISTROS V2 PARA PERIODO {period}\n")
        f.write(f"Fecha de Ejecucion: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("="*60 + "\n")
        f.write(f"Consulta Externa (LIMPIA): {len(final_consulta)} filas\n")
        f.write(f"Emergencia (LIMPIA): {len(final_emergencia)} filas\n")
        f.write(f"Hospitalización (LIMPIA): {len(final_hospitalizacion)} filas\n")
        f.write(f"Farmacia (LIMPIA): {len(final_farmacia)} filas\n")
        f.write(f"Estancias Pares E->H Retenidas (Caso A): {len(eh_groups)} pares\n")
        f.write(f"Registros de Paquete Retenidos: {len(excluded_package_emergencia) + len(excluded_package_hospitalizacion)} filas\n")
        f.write(f"Procedimientos Duplicados entre Fuentes Retenidos: {len(retained_duplicates_sources)} filas\n")
        f.write(f"Procedimientos Duplicados de Origen (Informativos): {len(duplicates_origin_list)} filas\n")
        f.write(f"Emergencias Huérfanas de Transferencia: {len(huerfanas_list)} filas\n")
        
    print("controles_integridad.txt summary written.")
    
    # 8. Generate metricas.json in 03_INFORMATIVOS/
    canonico = "CPT" if month < 10 else "SIGESAPOL"
    
    # Query duplicate stats
    cur.execute("""
        SELECT COUNT(*), COALESCE(SUM(sig.sp_valorizacion_calculada), 0)
        FROM temp_cpt_procedimientos_unificado cpt
        JOIN temp_sigesapol_procedimientos sig
          ON sig.sp_numero_documento_paciente = cpt.numero_documento_paciente
         AND sig.sp_fecha_atencion::date = cpt.fecha_atencion::date
         AND sig.sp_codigo_procedimiento = cpt.codigo_procedimiento
         AND (   (sig.tipo_procedimiento = 1
                  AND sig.sp_numero_documento_responsable = cpt.numero_documento_responsable)
              OR (sig.tipo_procedimiento IN (2, 3)
                  AND sig.sp_suma_cantidad = cpt.suma_cantidad_registro) );
    """)
    dup_cnt, dup_val = cur.fetchone()
    
    metrics = {
        "periodo": period,
        "fuente_canonica": canonico,
        "volumenes_raw": {
            "consulta": len(consulta_raw),
            "emergencia": len(emergencia_raw),
            "hospitalizacion": len(hospitalizacion_raw),
            "farmacia": len(farmacia_raw)
        },
        "volumenes_tramas": {
            "trama_consulta_externa": len(final_consulta),
            "trama_emergencia": len(final_emergencia),
            "trama_hospitalizacion": len(final_hospitalizacion),
            "trama_farmacia": len(final_farmacia)
        },
        "deduplicacion": {
            "duplicados_ciertos": int(dup_cnt),
            "monto_evitado_doble_cobro": float(dup_val)
        },
        "observaciones": {
            "solapamientos_estancias": len(eh_groups),
            "duplicados_fuentes": len(retained_duplicates_sources) // 2,
            "duplicados_origen": len(duplicates_origin_list) // 2,
            "transiciones_huerfanas": len(huerfanas_list)
        },
        "conservacion": conservacion
    }
    
    with open(os.path.join(infos_dir, "metricas.json"), 'w', encoding='utf-8') as f:
        json.dump(metrics, f, indent=4)
        
    print("metricas.json successfully written.")
    
    cur.close()
    conn.close()

if __name__ == "__main__":
    main()
